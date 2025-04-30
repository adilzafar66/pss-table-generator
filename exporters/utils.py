from consts import styles
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from consts.common import SUBHEAD_ROW, CONFIG_MAP, HEADER_ROW


def apply_row_format(sheet: Worksheet, start_row):
    """
    Helper method to apply formatting to data rows.

    :param Worksheet sheet: The worksheet object.
    :param int start_row: Row index to start formatting at.
    """
    for i in range(start_row, sheet.max_row + 1):
        for cell in sheet[i]:
            fill = get_row_cell_fill(cell)
            apply_cell_format(cell, styles.FONT_ENTRIES, styles.ALIGNMENT, styles.BORDER_ALL, fill)


def get_sorted_configs(data: dict) -> list:
    """
    Retrieves and sorts switching configurations based on passed data.

    :return: A list of sorted configurations.
    :rtype: list
    """
    configs = set()
    random_key = next(iter(data))
    for _id, entry in data[random_key].items():
        random_dict = next((v for v in entry.values() if isinstance(v, dict)), {})
        configs.update(random_dict.keys())
    return rearrange_list(list(configs), list(CONFIG_MAP.keys()))


def rearrange_list(target_list: list, reference_list: list) -> list:
    """
    Rearranges a target list based on the order of a reference list.

    :param list target_list: The list to be rearranged.
    :param list reference_list: The list that provides the desired order.
    :return: A new list with elements ordered according to the reference list.
    :rtype: list
    """
    reference_dict = {element: index for index, element in enumerate(reference_list)}
    in_reference = [element for element in target_list if element in reference_dict]
    not_in_reference = [element for element in target_list if element not in reference_dict]
    in_reference_sorted = sorted(in_reference, key=lambda x: reference_dict[x])
    return in_reference_sorted + not_in_reference


def set_const_cols_header(sheet: Worksheet, header: str, const_cols: list):
    """
    Helper method to set the main header for a sheet.

    :param Worksheet sheet: The worksheet object.
    :param str header: The header name to populate the cells with.
    :param list const_cols: List of constant columns.
    """
    sheet.cell(HEADER_ROW, 1).value = header
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(const_cols))
    set_cols_subheads(sheet, const_cols, start_col=1)


def set_var_cols_headers(sheet: Worksheet, configs: list, var_cols: list, col_prefix: str,
                         const_cols_buff: int, var_cols_buff: int):
    """
    Helper method to set dynamic headers for configurations.

    :param Worksheet sheet: The worksheet object.
    :param list configs: List of configurations to create columns for.
    :param list var_cols: List of variable columns.
    :param str col_prefix: Prefix for dynamic column headers.
    :param int const_cols_buff: Offset for constant columns.
    :param int var_cols_buff: Offset for variable columns.
    """
    for i, config in enumerate(configs):
        col_name = f'{col_prefix}: {CONFIG_MAP.get(config, config)}'
        col_index = const_cols_buff + var_cols_buff * i
        sheet.cell(HEADER_ROW, col_index).value = col_name
        sheet.merge_cells(start_row=1, start_column=col_index, end_row=1,
                          end_column=col_index + len(var_cols) - 1)
        set_cols_subheads(sheet, var_cols, start_col=col_index)


def set_end_cols_headers(sheet: Worksheet, header: str | None, var_cols: list, const_cols_buff: int, offset: int):
    """
    Helper method to set end headers for a sheet.

    :param Worksheet sheet: The worksheet object.
    :param str | None header: The header name to populate the cells with.
    :param list var_cols: List of variable columns.
    :param int const_cols_buff: Offset for constant columns.
    :param int offset: Offset for the end headers.
    """
    if not header:
        return
    last_col_index = const_cols_buff + offset
    sheet.cell(HEADER_ROW, last_col_index).value = header
    sheet.merge_cells(start_row=1, start_column=last_col_index, end_row=1,
                      end_column=last_col_index + len(var_cols) - 1)
    set_cols_subheads(sheet, var_cols, start_col=last_col_index)


def set_cols_subheads(sheet: Worksheet, columns: list, start_col: int):
    """
    Helper method to set subheads for a worksheet.

    :param Worksheet sheet: The worksheet object.
    :param list columns: List of subhead column names.
    :param int start_col: The starting column index for subhead.
    """
    for i, col_name in enumerate(columns):
        sheet.cell(SUBHEAD_ROW, start_col + i).value = col_name


def apply_cell_format(cell: Cell, font, align, border, fill=None):
    """
    Applies formatting to a single cell.

    :param Cell cell: The cell to apply formatting to.
    :param font: Font style for the cell.
    :param align: Alignment for the cell.
    :param border: Border style for the cell.
    :param fill: Fill pattern for the cell.
    """
    cell.font = font
    cell.alignment = align
    cell.border = border
    if fill is not None:
        cell.fill = fill


def get_row_cell_fill(cell: Cell):
    """
    Applies fill to a single row cell.

    :param Cell cell: The cell to apply fill to.
    """
    if cell.row % 2 == 0 and not cell.fill.patternType:
        return styles.FILL_ROW_BLUE


def set_column_widths(sheet: Worksheet, const_cols_len: int, var_cols_len: int, col_width: float):
    """
    Helper method to set column widths.

    :param Worksheet sheet: The worksheet object.
    :param int const_cols_len: Number of constant columns.
    :param int var_cols_len: Number of variable columns.
    :param float col_width: Width of the columns.
    """
    for i in range(1, sheet.max_column + 1):
        if i % (var_cols_len + 1) == (const_cols_len + 1) % (var_cols_len + 1) and i > const_cols_len:
            sheet.column_dimensions[get_column_letter(i)].width = styles.WIDTH_BUFFER
            for column in sheet.iter_cols(i, i):
                for cell in column:
                    cell.fill = styles.FILL_ROW_BLANK
                    cell.border = styles.BORDER_VERTICAL
        else:
            sheet.column_dimensions[get_column_letter(i)].width = col_width


def format_numbers(sheet: Worksheet):
    """
    Formats cells with numerical values.

    :param Worksheet sheet: The worksheet object.
    """
    for i in range(1, sheet.max_column + 1):
        for column in sheet.iter_cols(i, i):
            for cell in column:
                cell.number_format = styles.NUMBER_FORMAT
