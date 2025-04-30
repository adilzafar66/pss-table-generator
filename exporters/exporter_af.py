from consts import styles
from openpyxl import Workbook
from consts.common import SUBHEAD_ROW
from consts.sheets import WS_ARC_FLASH
from exporters.exporter import Exporter
from consts.columns import AF_SI_CONST_COLS, AF_CONST_COLS


class ArcFlashExporter(Exporter):
    """
    A class to export Arc Flash data to an Excel workbook using the openpyxl library.

    Attributes:
        wb (Workbook): The Excel workbook.
        ws (Worksheet): The active worksheet in the workbook.
    """

    def __init__(self):
        """
        Initializes a new instance of the ArcFlashExporter class,
        creating a new workbook and setting up the active worksheet.
        """
        super().__init__([WS_ARC_FLASH], [AF_CONST_COLS, None])
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = WS_ARC_FLASH

    def add_data(self, af_data: dict):
        """
        Adds data rows to the worksheet.

        :param dict af_data: A dictionary where the keys are row headers and values
                             are lists of data corresponding to each column.
        """
        for i, (key, data) in enumerate(af_data.items()):
            self.ws.append([key] + data)

    def create_headers(self, use_si_units: bool = False, **args):
        """
        Creates the header row in the worksheet using predefined column names
        and applies header-specific formatting.

        :param bool use_si_units: A flag to determine whether to use column headings with SI units.
        """
        column_names = AF_SI_CONST_COLS if use_si_units else AF_CONST_COLS
        for index, col_name in enumerate(column_names):
            cell = self.ws.cell(1, index + 1)
            cell.value = col_name
            cell.fill = styles.FILL_HEADER
            cell.font = styles.FONT_HEADER
            cell.alignment = styles.ALIGNMENT
            cell.border = styles.BORDER_ALL

    def highlight_high_energy(self, max_energy: float, crit_energy: float):
        """
        Highlights cells in the 'Total Energy' column that fall within specified
        energy thresholds by applying different fill colors.

        :param float max_energy: The maximum energy threshold for applying low energy highlighting.
        :param float crit_energy: The critical energy threshold for applying high energy highlighting.
        """
        energy_col = self._get_col_index(0, 'Total Energy') + 1
        for column in self.ws.iter_cols(energy_col, energy_col, min_row=SUBHEAD_ROW):
            for cell in column:
                if max_energy < cell.value < crit_energy:
                    cell.fill = styles.FILL_ROW_ORANGE
                elif cell.value > crit_energy:
                    cell.fill = styles.FILL_ROW_RED
