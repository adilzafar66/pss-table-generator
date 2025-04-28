from pathlib import Path
from consts import styles
from exporters import utils
from openpyxl.workbook import Workbook
from consts.common import CONFIG_MAP, SUBHEAD_ROW, HEADER_ROW


class Exporter:

    def __init__(self, sheet_names: list[str], header_names: list[str]):
        """
        Initializes a new instance of the Exporter class,
        creating a new workbook and setting up the worksheets.
        """
        self.ansi_data = {}
        self.iec_data = {}
        self.wb = Workbook()
        self._initialize_sheets(sheet_names)
        self.header_names = header_names

    def _initialize_sheets(self, ws_names):
        """
        Initializes and names the worksheets.
        """

        self.wb.active.title = ws_names[0]
        for ws_name in ws_names[1:]:
            self.wb.create_sheet(ws_name)

    def _get_col_index(self, sheet_index: int, heading: str):
        """
        Gets the column index for a given column heading in a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param str heading: The heading of the column to find.
        :return: The index of the column with the specified heading.
        :rtype: int
        """
        sheet = self.wb.worksheets[sheet_index]
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value and heading in col[0].value:
                return col[0].column - 1

    def set_ansi_data(self, ansi_data: dict):
        """
        Sets ANSI data and updates the configurations.

        :param dict ansi_data: A dictionary containing ANSI data.
        """
        self.ansi_data = ansi_data

    def set_iec_data(self, iec_data: dict):
        """
        Sets IEC data and updates the configurations.

        :param dict iec_data: A dictionary containing IEC data.
        """
        self.iec_data = iec_data

    def create_headers(self, sheet_index: int, const_cols: list, var_cols: list, col_prefix: str):
        """
        Creates and merges headers for a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param list const_cols: List of constant columns.
        :param list var_cols: List of variable columns.
        :param str col_prefix: Prefix for dynamic column headers.
        """
        var_cols_buff = len(var_cols) + 1
        const_cols_buff = len(const_cols) + 2
        sheet = self.wb.worksheets[sheet_index]
        configs = utils.get_sorted_configs(self.ansi_data)
        const_cols_header, end_cols_header = self.header_names
        utils.set_const_cols_header(sheet, const_cols_header, const_cols)
        utils.set_var_cols_headers(sheet, configs, var_cols, col_prefix, const_cols_buff, var_cols_buff)
        utils.set_end_cols_headers(sheet, end_cols_header, var_cols, const_cols_buff, len(configs) * var_cols_buff)

    def format_headers(self, sheet_index: int):
        """
        Applies formatting to the header rows of a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        """
        sheet = self.wb.worksheets[sheet_index]
        for row in sheet.iter_rows(HEADER_ROW, SUBHEAD_ROW):
            for cell in row:
                utils.apply_cell_format(cell, styles.FONT_HEADER, styles.ALIGNMENT,
                                        styles.BORDER_ALL, styles.FILL_HEADER)

    def format_sheet(self, sheet_index: int, const_cols_len: int, var_cols_len: int, col_width: float):
        """
        Applies formatting to the entire sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param int const_cols_len: Number of constant columns.
        :param int var_cols_len: Number of variable columns.
        :param float col_width: Width of the columns.
        """
        sheet = self.wb.worksheets[sheet_index]
        utils.apply_row_format(sheet, SUBHEAD_ROW + 1)
        utils.set_column_widths(sheet, const_cols_len, var_cols_len, col_width)
        utils.format_numbers(sheet)

    def save_workbook(self, wb_path: Path):
        """
        Saves the workbook to a specified filename.

        :param Path wb_path: The filename to save the workbook as.
        """
        self.wb.save(wb_path)
        self.wb.close()