from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEIGHT_ROW = 15
WIDTH_COL_LRG = 20
WIDTH_COL_SML = 18
WIDTH_BUFFER = 2
NUMBER_FORMAT = '0.00#'
FONT_ENTRIES = Font(name='Calibri', bold=False, size=8)
FONT_HEADER = Font(name='Calibri', bold=True, size=8, color='FFFFFF')
FILL_HEADER = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
FILL_ROW_BLUE = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
FILL_ROW_RED = PatternFill(start_color='FFABB3', end_color='FFABB3', fill_type='solid')
FILL_ROW_ORANGE = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
FILL_ROW_YELLOW = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')
FILL_ROW_BLANK = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
ALIGNMENT = Alignment(horizontal='center', vertical='center')
BORDER_ALL = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
BORDER_VERTICAL = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style=None),
                         bottom=Side(style=None))