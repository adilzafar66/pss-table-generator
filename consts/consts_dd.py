from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DD_STUDY_TAG = 'DD'
DD_STUDY_CASE = 'SC'
DD_STUDY_CASE_IEC = 'SC_IEC'
DD_STUDY_MODE = 'ANSI DEVICE DUTY'
DD_STUDY_MODE_IEC = 'IEC DEVICE DUTY'
DD_STUDY_MODE_IEC_1P = 'IEC_1PHASE_DEVICE_DUTY'
DD_STUDY_MODE_1P = 'ANSI_1PHASE_DEVICE_DUTY'
SYSTEM = 'Network Analysis'
PRESENTATION = 'OLV1'

MODES = {
    'Momentary': 'MOM',
    'Interrupt': 'INT'
}

CONFIG_MAP = {
    'PRES': 'Present',
    'PRES-A': 'Present A',
    'PRES-B': 'Present B',
    'ULT': 'Ultimate',
    'ULT-A': 'Ultimate A',
    'ULT-B': 'Ultimate B',
    'GEN': 'Generator',
    'MAX': 'Maximum',
    'MIN': 'Minimum'
}

CONFIG_MAP_INV = {
    'Normal': 'PRES',
    'Ultimate': 'ULT',
    'Generator': 'GEN'
}

TOP_COLS = [
    'Device',
    'Device Capability'
]

SPECS = {
    'header_row': 1,
    'subheader_row': 2,
    'font_data': Font(name='Calibri', bold=False, size=8),
    'font_header': Font(name='Calibri', bold=True, size=8, color='FFFFFF'),
    'font_cmp': Font(name='Calibri', bold=False, size=8, color='FFFFFF'),
    'fill_header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
    'fill_alt': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
    'fill_cmp': PatternFill(start_color='FFABB3', end_color='FFABB3', fill_type='solid'),
    'fill_sr': PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid'),
    'fill_blank': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'align': Alignment(horizontal='center', vertical='center'),
    'row_height': 15,
    'col_width_lrg': 21,
    'col_width_sml': 18,
    'buffer_width': 2,
    'border': Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin')),
    'border_none': Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style=None),
                          bottom=Side(style=None))
}

mom_const_cols = ['ID', 'Nominal kV', 'Type']
mom_alt_cols = ['Symm. kA rms', 'Asymm. kA rms']
int_const_cols = ['ID', 'Voltage', 'Bus', 'Device']
int_alt_cols = ['Int. Adj. Symm. kA']
int_iec_alt_cols = ['Ib Sym. kA', 'Ib Asym. kA']

spec_keys_mom = {
    'Type': 'Type',
    'fault_head': ['Sym', 'Asym'],
    'cap_fault_head': ['CapSym', 'CapAsym']
}

spec_keys_int = {
    'Type': 'Bus',
    'fault_head': ['AdjSym'],
    'cap_fault_head': ['CapAdjSym']
}

spec_keys_iec_int = {
    'Type': 'Bus',
    'fault_head': ['LbSym', 'LbAsym'],
    'cap_fault_head': ['CapLbSym', 'CapLbAsym']
}

ANSI_EXT = 'SA1S'
ANSI_SP_EXT = 'SA4S'
IEC_EXT = 'SI1S'
IEC_SP_EXT = 'SI4S'
MULTIPLIER = 1.246
MV_SWITCHGEAR_MULTIPLIER = 1.55
LV_SWITCHGEAR_MULTIPLIER = 1.33
FILE_NAME_SUFFIX = 'Device Duty Report.xlsx'
