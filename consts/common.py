from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROUND_DIGITS = 2
FILE_NAME_SUFFIX = 'Arc Flash Report.xlsx'

SYSTEM = 'Network Analysis'
PRESENTATION = 'OLV1'
BASE_REVISION = 'Base'
SC_STUDY_TAG = 'SC'
SC_STUDY_CASE = 'SC'
SC_STUDY_MODE = 'ANSI ALL FAULT MOMENTARY'

DEFAULT_SW_CONFIGS = [
    'Normal',
    'Present',
    'PRES',
    'Ultimate',
    'ULT',
    'Generator',
    'GEN'
]
HTTP = 'http'
HTTPS = 'https'

TYPE_MAP = {
    'Fuse': 'FUSE',
    'SPDT Switch': 'DOUBLESWITCH',
    'SPST Switch': 'SINGLESWITCH',
    'Molded Case': 'LVCB'
}

SHORT_CIRCUIT_SHEET = 'Short Circuit'
SEQUENCE_IMP_SHEET = 'Sequence Impedance'

CONFIG_MAP = {
    'PRES': 'Present',
    'STBY': 'Standby',
    'ULT': 'Ultimate',
    'GEN': 'Generator',
    'MAX': 'Maximum',
    'MIN': 'Minimum',
}

INV_CONFIG_MAP = {
    'Present': 'PRES',
    'Normal': 'PRES',
    'Standby': 'STBY',
    'Ultimate': 'ULT',
    'Generator': 'GEN'
}

SPECS = {
    'header_row': 1,
    'subheader_row': 2,
    'font_data': Font(name='Calibri', bold=False, size=8),
    'font_header': Font(name='Calibri', bold=True, size=8, color='FFFFFF'),
    'font_cmp': Font(name='Calibri', bold=False, size=8, color='FFFFFF'),
    'fill_header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
    'fill_alt': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
    'fill_cmp': PatternFill(start_color='FFABB3', end_color='FFABB3', fill_type='solid'),
    'fill_sr': PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid'),
    'fill_blank': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'align': Alignment(horizontal='center', vertical='center'),
    'row_height': 15,
    'col_width_lrg': 21,
    'col_width_sml': 18,
    'buffer_width': 2,
    'border': Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin')),
    'border_none': Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style=None),
                          bottom=Side(style=None))
}