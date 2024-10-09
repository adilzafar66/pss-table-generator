from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AF_STUDY_TAG = 'AF'
AF_STUDY_MODE = 'ANSI ARC-FLASH'
VCB_CONFIG = 'VCB'
VCBB_CONFIG = 'VCBB'
ROUND_DIGITS = 2
FT_M_MULTIPLIER = 0.3048
FILE_NAME_SUFFIX = 'Arc Flash Report.xlsx'

indices = {
    'ecf': 2,
    'lab': 4,
    'rab': 5,
    'rep': 6,
    'con': 7,
    'ie': 8,
    'afb': 9,
    'fct': 10,
    'la_var': 12
}

CONFIG_MAP = {
    'Normal': 'PRES',
    'GEN': 'GEN',
    'Normal_NM': 'PRES_NM',
    'GEN_NM': 'GEN_NM',
    'Normal-NM': 'PRES_NM',
    'GEN-NM': 'GEN_NM'
}

COLUMN_NAMES = [
    'ID',
    'kV',
    'Type',
    'Electrode Configuration',
    'Working Distance LL (in)',
    'LAB to Fixed Part (ft-in)',
    'RAB (ft-in)',
    'Output Rpt.',
    'Configuration',
    'Total Energy (cal/cm²)',
    'AFB (ft-in)',
    'Final FCT (sec)',
    'Source PD ID',
    '% Ia Variation',
    "Total Ia'' (kA)",
    'Source PD Ia at FCT (kA)',
    'Total Ibf at FCT (kA)',
    'Source PD Ibf at FCT (kA)'
]

COLUMN_NAMES_SI = [
    'ID',
    'kV',
    'Type',
    'Electrode Configuration',
    'Working Distance LL (m)',
    'LAB to Fixed Part (m)',
    'RAB (m)',
    'Output Rpt.',
    'Configuration',
    'Total Energy (cal/cm²)',
    'AFB (m)',
    'Final FCT (sec)',
    'Source PD ID',
    '% Ia Variation',
    "Total Ia'' (kA)",
    'Source PD Ia at FCT (kA)',
    'Total Ibf at FCT (kA)',
    'Source PD Ibf at FCT (kA)'
]

SPECS = {
    'header_row': 1,
    'subheader_row': 2,
    'font_data': Font(name='Calibri', bold=False, size=8),
    'font_header': Font(name='Calibri', bold=True, size=8, color='FFFFFF'),
    'font_cmp': Font(name='Calibri', bold=False, size=8, color='FFFFFF'),
    'fill_header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
    'fill_alt': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
    'fill_cmp_high': PatternFill(start_color='FFABB3', end_color='FFABB3', fill_type='solid'),
    'fill_cmp_low': PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),
    'fill_sr': PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid'),
    'fill_blank': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
    'align': Alignment(horizontal='center', vertical='center'),
    'row_height': 15,
    'col_width_lrg': 20,
    'col_width_sml': 18,
    'buffer_width': 2,
    'border': Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin')),
    'border_none': Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style=None),
                          bottom=Side(style=None))
}


