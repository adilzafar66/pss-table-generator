from pathlib import Path
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from consts.consts_dd import SPECS, TOP_COLS, CONFIG_MAP

HEADER_ROW = SPECS['header_row']
SUBHEADER_ROW = SPECS['subheader_row']
FILL_HEADER = SPECS['fill_header']
FONT_HEADER = SPECS['font_header']
ALIGN = SPECS['align']
BORDER = SPECS['border']
FILL_ALT = SPECS['fill_alt']
FILL_BLANK = SPECS['fill_blank']
BORDER_NONE = SPECS['border_none']
FONT_DATA = SPECS['font_data']
BUFFER_WIDTH = SPECS['buffer_width']
NUMBER_FORMAT = '0.00#'

ANSI_MOM_SHEET = 'ANSI Momentary Table'
ANSI_INT_SHEET = 'ANSI Interrupting Table'
IEC_INT_SHEET = 'IEC Interrupting Table'


class DeviceDutyExporter:
    """
    A class to export Device Duty data to an Excel workbook using the openpyxl library.

    Attributes:
        ansi_data (dict): Data specific to ANSI standards.
        iec_data (dict): Data specific to IEC standards.
        wb (Workbook): The Excel workbook.
        ws_mom (Worksheet): The worksheet for ANSI Momentary Table.
        ws_int (Worksheet): The worksheet for ANSI Interrupting Table.
        ws_iec (Worksheet): The worksheet for IEC Interrupting Table.
        configs (list): Sorted list of configurations used in headers.
    """

    def __init__(self):
        """
        Initializes a new instance of the DeviceDutyExporter class,
        creating a new workbook and setting up the worksheets.
        """
        self.ansi_data = {}
        self.iec_data = {}
        self.wb = Workbook()
        self._initialize_sheets()
        self.configs = []

    def _initialize_sheets(self):
        """
        Initializes and names the workbook sheets.
        """
        self.wb.active.title = ANSI_MOM_SHEET
        self.wb.create_sheet(ANSI_INT_SHEET)
        self.wb.create_sheet(IEC_INT_SHEET)
        self.ws_mom, self.ws_int, self.ws_iec = self.wb.worksheets

    def set_ansi_data(self, ansi_data: dict):
        """
        Sets ANSI data and updates the configurations.

        :param dict ansi_data: A dictionary containing ANSI data.
        """
        self.ansi_data = ansi_data
        self.configs = self._get_sorted_configs()

    def set_iec_data(self, iec_data: dict):
        """
        Sets IEC data and updates the configurations.

        :param dict iec_data: A dictionary containing IEC data.
        """
        self.iec_data = iec_data
        self.configs = self._get_sorted_configs()

    def _get_sorted_configs(self) -> list:
        """
        Retrieves and sorts configurations based on ANSI data.

        :return: A list of sorted configurations.
        :rtype: list
        """
        configs = set()
        for _id, details in self.ansi_data['MOM'].items():
            for config in details['Sym']:
                configs.add(config)
        return self.rearrange_list(list(configs), list(CONFIG_MAP.keys()))

    def _get_col_index(self, sheet_index: int, heading: str):
        """
        Gets the column index for a given column heading in a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param str heading: The heading of the column to find.
        :return: The index of the column with the specified heading.
        :rtype: int
        """
        sheet = self.wb.worksheets[sheet_index]
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value and heading in col[0].value:
                return col[0].column - 1

    @staticmethod
    def rearrange_list(target_list: list, reference_list: list) -> list:
        """
        Rearranges a target list based on the order of a reference list.

        :param list target_list: The list to be rearranged.
        :param list reference_list: The list that provides the desired order.
        :return: A new list with elements ordered according to the reference list.
        :rtype: list
        """
        reference_dict = {element: index for index, element in enumerate(reference_list)}
        in_reference = [element for element in target_list if element in reference_dict]
        not_in_reference = [element for element in target_list if element not in reference_dict]
        in_reference_sorted = sorted(in_reference, key=lambda x: reference_dict[x])
        return in_reference_sorted + not_in_reference

    @staticmethod
    def highlight_high_duty(cell: Cell, fault_val: float, cap_val: float):
        """
        Highlights cells where the fault value exceeds the capability value.

        :param Cell cell: The cell to apply formatting to.
        :param float fault_val: The fault value to compare.
        :param float cap_val: The capability value to compare against.
        """
        if fault_val > cap_val:
            cell.fill = SPECS['fill_cmp']

    @staticmethod
    def highlight_series_rated(cell, is_series_rated: bool):
        """
        Highlights cells that are series rated.

        :param Cell cell: The cell to apply formatting to.
        :param bool is_series_rated: Indicates whether the cell should be series rated.
        """
        if is_series_rated:
            cell.fill = SPECS['fill_sr']

    def create_headers(self, sheet_index: int, const_cols: list, alt_cols: list, col_prefix: str):
        """
        Creates and merges headers for a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param list const_cols: List of constant columns.
        :param list alt_cols: List of alternating columns.
        :param str col_prefix: Prefix for dynamic column headers.
        """
        alt_cols_buff = len(alt_cols) + 1
        const_cols_buff = len(const_cols) + 2
        sheet = self.wb.worksheets[sheet_index]
        self._set_main_header(sheet, const_cols)
        self._set_subheaders(sheet, const_cols, start_col=1)
        self._set_dynamic_headers(sheet, alt_cols, col_prefix, const_cols_buff, alt_cols_buff)
        self._set_end_headers(sheet, alt_cols, const_cols_buff, len(self.configs) * alt_cols_buff)

    @staticmethod
    def _set_main_header(sheet: Worksheet, const_cols: list):
        """
        Helper method to set the main header for a sheet.

        :param Worksheet sheet: The worksheet object.
        :param list const_cols: List of constant columns.
        """
        sheet.cell(HEADER_ROW, 1).value = TOP_COLS[0]
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(const_cols))

    @staticmethod
    def _set_subheaders(sheet: Worksheet, columns: list, start_col: int):
        """
        Helper method to set subheaders for a sheet.

        :param Worksheet sheet: The worksheet object.
        :param list columns: List of subheader column names.
        :param int start_col: The starting column index for subheaders.
        """
        for i, col_name in enumerate(columns):
            sheet.cell(SUBHEADER_ROW, start_col + i).value = col_name

    def _set_dynamic_headers(self, sheet: Worksheet, alt_cols: list, col_prefix: str,
                             const_cols_buff: int, alt_cols_buff: int):
        """
        Helper method to set dynamic headers for configurations.

        :param Worksheet sheet: The worksheet object.
        :param list alt_cols: List of alternating columns.
        :param str col_prefix: Prefix for dynamic column headers.
        :param int const_cols_buff: Offset for constant columns.
        :param int alt_cols_buff: Offset for alternating columns.
        """
        for i, config in enumerate(self.configs):
            col_name = f'{col_prefix}: {CONFIG_MAP.get(config, config)}'
            col_index = const_cols_buff + alt_cols_buff * i
            sheet.cell(HEADER_ROW, col_index).value = col_name
            sheet.merge_cells(start_row=1, start_column=col_index, end_row=1,
                              end_column=col_index + len(alt_cols) - 1)
            self._set_subheaders(sheet, alt_cols, start_col=col_index)

    def _set_end_headers(self, sheet: Worksheet, alt_cols: list, const_cols_buff: int, offset: int):
        """
        Helper method to set end headers for a sheet.

        :param Worksheet sheet: The worksheet object.
        :param list alt_cols: List of alternating columns.
        :param int const_cols_buff: Offset for constant columns.
        :param int offset: Offset for the end headers.
        """
        last_col_index = const_cols_buff + offset
        sheet.cell(HEADER_ROW, last_col_index).value = TOP_COLS[1]
        sheet.merge_cells(start_row=1, start_column=last_col_index, end_row=1,
                          end_column=last_col_index + len(alt_cols) - 1)
        self._set_subheaders(sheet, alt_cols, start_col=last_col_index)

    def format_headers(self, sheet_index: int):
        """
        Applies formatting to the header rows of a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        """
        sheet = self.wb.worksheets[sheet_index]
        for row in sheet.iter_rows(HEADER_ROW, SUBHEADER_ROW):
            for cell in row:
                self._apply_cell_format(cell, FONT_HEADER, ALIGN, BORDER, FILL_HEADER)

    @staticmethod
    def _apply_cell_format(cell: Cell, font, align, border, fill=None):
        """
        Applies formatting to a single cell.

        :param Cell cell: The cell to apply formatting to.
        :param font: Font style for the cell.
        :param align: Alignment for the cell.
        :param border: Border style for the cell.
        :param fill: Fill pattern for the cell.
        """
        cell.font = font
        cell.alignment = align
        cell.border = border
        if fill is not None:
            cell.fill = fill

    def format_sheet(self, sheet_index: int, const_cols_len: int, alt_cols_len: int, col_width: float):
        """
        Applies formatting to the entire sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param int const_cols_len: Number of constant columns.
        :param int alt_cols_len: Number of alternating columns.
        :param float col_width: Width of the columns.
        """
        sheet = self.wb.worksheets[sheet_index]
        self._apply_row_format(sheet)
        self._set_column_widths(sheet, const_cols_len, alt_cols_len, col_width)
        self._format_numbers(sheet)

    def _apply_row_format(self, sheet: Worksheet):
        """
        Helper method to apply formatting to data rows.

        :param Worksheet sheet: The worksheet object.
        """
        for i in range(3, sheet.max_row + 1):
            for cell in sheet[i]:
                fill = ((cell.row % 2 == 0 and not cell.fill.patternType) and FILL_ALT) or None
                self._apply_cell_format(cell, FONT_DATA, ALIGN, BORDER, fill)

    @staticmethod
    def _set_column_widths(sheet: Worksheet, const_cols_len: int, alt_cols_len: int, col_width: float):
        """
        Helper method to set column widths.

        :param Worksheet sheet: The worksheet object.
        :param int const_cols_len: Number of constant columns.
        :param int alt_cols_len: Number of alternating columns.
        :param float col_width: Width of the columns.
        """
        for i in range(1, sheet.max_column + 1):
            if i % (alt_cols_len + 1) == (const_cols_len + 1) % (alt_cols_len + 1) and i > const_cols_len:
                sheet.column_dimensions[get_column_letter(i)].width = BUFFER_WIDTH
                for column in sheet.iter_cols(i, i):
                    for cell in column:
                        cell.fill = FILL_BLANK
                        cell.border = BORDER_NONE
            else:
                sheet.column_dimensions[get_column_letter(i)].width = col_width

    @staticmethod
    def _format_numbers(sheet: Worksheet):
        """
        Formats cells with numerical values.

        :param Worksheet sheet: The worksheet object.
        """
        for i in range(1, sheet.max_column + 1):
            for column in sheet.iter_cols(i, i):
                for cell in column:
                    cell.number_format = NUMBER_FORMAT

    def insert_data(self, sheet_index: int, data_key: str, spec_keys: dict, dataset: str = 'ansi'):
        """
        Inserts data into a specified sheet.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param str data_key: Key for the data section to insert.
        :param dict spec_keys: Dictionary of specification keys used to extract data.
        :param str dataset: The dataset to use ('ansi' or 'iec'). Default is 'ansi'.
        """
        data = self.ansi_data if dataset == 'ansi' else self.iec_data
        sheet = self.wb.worksheets[sheet_index]
        start_row = SPECS['subheader_row'] + 1
        ids = list(data[data_key].keys())
        values = list(data[data_key].values())
        for i in range(len(data[data_key])):
            row = sheet[start_row + i]
            row[0].value = ids[i].strip() + '*' if values[i].get('Assumed', False) else ids[i].strip()
            row[1].value = round(values[i]['Voltage'], 3)
            row[2].value = values[i].get(spec_keys['Type'], '').strip()
            row[3].value = values[i].get('Device', '').strip()
            fault_vals = [values[i][k] for k in spec_keys['fault_head']]
            cap_vals = [values[i][k] for k in spec_keys['cap_fault_head']]
            for j in range(len(fault_vals)):
                self.insert_fault_data(sheet_index, row, fault_vals[j], cap_vals[j], offset=j)
            last_col_index = self._get_col_index(sheet_index, TOP_COLS[1])
            for j in range(len(cap_vals)):
                row[last_col_index + j].value = round(cap_vals[j], 2) or '--'
                # if values[i].get('Assumed') and round(cap_vals[j], 2):
                #     row[last_col_index + j].value = str(row[last_col_index + j].value) + '*'
                self.highlight_series_rated(row[last_col_index + j], values[i].get('SeriesRated', False))

    def insert_fault_data(self, sheet_index: int, row: tuple[Cell, ...], values: dict, cap_val: float, offset: int = 0):
        """
        Inserts fault data into specified cells in a row.

        :param int sheet_index: Index of the worksheet in the workbook.
        :param tuple[Cell, ...] row: The row to insert data into.
        :param dict values: A dictionary of fault values by configuration.
        :param float cap_val: The capability value to compare against.
        :param int offset: Offset for column insertion, default is 0.
        """
        for config, fault_val in values.items():
            heading_key = CONFIG_MAP.get(config, config)
            col_index = self._get_col_index(sheet_index, heading_key)
            if col_index:
                cell = row[col_index + offset]
                cell.value = round(fault_val, 2) or '--'
                self.highlight_high_duty(cell, fault_val, cap_val)

    def save_workbook(self, wb_path: Path):
        """
        Saves the workbook to a specified filename.

        :param Path wb_path: The filename to save the workbook as.
        """
        self.wb.save(wb_path)
        self.wb.close()
