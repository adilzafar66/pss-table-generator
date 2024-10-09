from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from consts.consts_af import COLUMN_NAMES, SPECS, COLUMN_NAMES_SI


class ArcFlashExporter:
    """
    A class to export Arc Flash data to an Excel workbook using the openpyxl library.

    Attributes:
        wb (Workbook): The Excel workbook.
        ws (Worksheet): The active worksheet in the workbook.
    """

    def __init__(self):
        """
        Initializes a new instance of the ArcFlashExporter class,
        creating a new workbook and setting up the active worksheet.
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'ANSI Arc Flash Table'

    def create_headers(self, use_si_units: bool = False):
        """
        Creates the header row in the worksheet using predefined column names
        and applies header-specific formatting.

        :param bool use_si_units: A flag to determine whether to use column headings with SI units.
        """
        column_names = COLUMN_NAMES_SI if use_si_units else COLUMN_NAMES
        for index, col_name in enumerate(column_names):
            cell = self.ws.cell(1, index + 1)
            cell.value = col_name
            cell.fill = SPECS['fill_header']
            cell.font = SPECS['font_header']
            cell.alignment = SPECS['align']
            cell.border = SPECS['border']

    def add_data(self, af_data: dict):
        """
        Adds data rows to the worksheet.

        :param dict af_data: A dictionary where the keys are row headers and values
                        are lists of data corresponding to each column.
        """
        for i, (key, data) in enumerate(af_data.items()):
            self.ws.append([key] + data)

    def format_sheet(self):
        """
        Applies overall formatting to the worksheet including row formatting,
        setting column widths, and formatting numbers.
        """
        self._apply_row_formatting()
        self._set_column_widths()
        self._format_numbers()

    def _apply_row_formatting(self):
        """
        Helper method that applies formatting to data rows, such as borders,
        fonts, alignments, and alternating row fill patterns.
        """
        for i in range(2, self.ws.max_row + 1):
            for cell in self.ws[i]:
                cell.border = SPECS['border']
                cell.font = SPECS['font_data']
                cell.alignment = SPECS['align']
                if cell.row % 2 != 0 and not cell.fill.patternType:
                    cell.fill = SPECS['fill_alt']

    def _set_column_widths(self):
        """
        Helper method that sets the widths of the columns in the worksheet
        based on predefined specifications.
        """
        for i in range(1, self.ws.max_column + 1):
            self.ws.column_dimensions[get_column_letter(i)].width = SPECS['col_width_lrg']

    def _format_numbers(self):
        """
        Helper method that applies number formatting to the cells in the worksheet,
        setting a specific numeric format for consistency.
        """
        for i in range(1, self.ws.max_column + 1):
            for column in self.ws.iter_cols(i, i):
                for cell in column:
                    cell.number_format = '0.00#'

    def highlight_high_energy(self, max_energy: float, crit_energy: float):
        """
        Highlights cells in the 'Total Energy' column that fall within specified
        energy thresholds by applying different fill colors.

        :param float max_energy: The maximum energy threshold for applying low energy highlighting.
        :param float crit_energy: The critical energy threshold for applying high energy highlighting.
        """
        energy_col = self.get_col_index('Total Energy')
        for column in self.ws.iter_cols(energy_col, energy_col, min_row=2):
            for cell in column:
                if max_energy < cell.value < crit_energy:
                    cell.fill = SPECS['fill_cmp_low']
                elif cell.value > crit_energy:
                    cell.fill = SPECS['fill_cmp_high']

    def get_col_index(self, heading: str) -> int:
        """
        Gets the column index for a given column heading.

        :param str heading: The heading of the column to find.
        :return: The index of the column with the specified heading.
        :rtype: int
        """
        for col in self.ws.iter_cols(1, self.ws.max_column):
            if col[0].value and heading in col[0].value:
                return col[0].column

    def save_workbook(self, wb_path: Path):
        """
        Saves the workbook to the specified path and closes it.

        :param Path wb_path: The file path where the workbook will be saved.
        """
        self.wb.save(wb_path)
        self.wb.close()
