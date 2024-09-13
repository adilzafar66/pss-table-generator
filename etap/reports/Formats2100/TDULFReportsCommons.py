#***********************
#
# Copyright (c) 2021-2023, Operation Technology, Inc.
#
# THIS PROGRAM IS CONFIDENTIAL AND PROPRIETARY TO OPERATION TECHNOLOGY, INC. 
# ANY USE OF THIS PROGRAM IS SUBJECT TO THE PROGRAM SOFTWARE LICENSE AGREEMENT, 
# EXCEPT THAT THE USER MAY MODIFY THE PROGRAM FOR ITS OWN USE. 
# HOWEVER, THE PROGRAM MAY NOT BE REPRODUCED, PUBLISHED, OR DISCLOSED TO OTHERS 
# WITHOUT THE PRIOR WRITTEN CONSENT OF OPERATION TECHNOLOGY, INC.
#
#***********************

import sqlite3 
from datetime import datetime 
from openpyxl import workbook, worksheet
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, _get_column_letter, \
    absolute_coordinate

def increase_column_by_one(coordinate):
     first_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][0]
     second_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][1]
     column_alphabet = coordinate[first_emp_index + 1:second_emp_index]
     if len(column_alphabet) == 1:
         if column_alphabet == 'Z':
             start_writing_column = 'AA'
         else:
             start_writing_column = chr(ord(column_alphabet)+1)

     elif len(column_alphabet) == 2:
         if column_alphabet == 'ZZ':
             start_writing_column = 'AAA'
         else:
             first_char_of_column_alphabet = column_alphabet[:len(column_alphabet)-1]
             second_char_of_column_alphabet = column_alphabet[len(column_alphabet)-1:]
             start_writing_column = first_char_of_column_alphabet + chr(ord(second_char_of_column_alphabet)+1)

     start_writing_coordinate = coordinate[:first_emp_index+1] + start_writing_column + coordinate[second_emp_index:]
     return start_writing_coordinate


def get_excelName_to_coordinates_dict(wb, db_column_name_list):
    column_names = len(wb.defined_names.definedName)    
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(wb.defined_names.definedName[i].name)           
    
    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]
    
    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix+db_column_name_list[count]] = db_column_name_list[count]
    
    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            excel_column_name = defined_names_for_other_sheet[i] #wb.defined_names.definedName[i].name
            actual_coordinates = next(wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+2
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue
    return excel_column_name_to_coordinates


def connect_to_project(project_path):
    # ....Connecting to SQLite DB....
    project = sqlite3.connect(project_path)
    cursor = project.cursor()

    # ....Get all the table names from report Database....
    all_tables_name = project.execute("SELECT name FROM sqlite_master WHERE type = 'table';")
    table_names_list = []
    for table_names in all_tables_name:
        table_names_list.append(table_names[0])

    return cursor, table_names_list


def get_updated_coordinates(coordinates):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) + 1
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates


def get_actual_day(int_val):
    int_to_day = {0: "Mon.", 1: "Tue.", 2: "Wed.", 3: "Thurs.", 4: "Fri.", 5: "Sat.", 6: "Sun."}
    value_day = int_to_day[int_val]
    return value_day


def patch_worksheet():
    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    	if not range_string and not all((start_row, start_column, end_row, end_column)):
    		raise ValueError()
    	elif not range_string:
    		range_string = '%s%s:%s%s' % (get_column_letter(start_column),
    									start_row,
    									get_column_letter(end_column),
    									end_row)
    	elif ":" not in range_string:
    		if COORD_RE.match(range_string):
    			return  # Single cell, do nothing
    		raise ValueError()
    	else:
    		range_string = range_string.replace('$', '')
    
    	if range_string not in self.merged_cells:
    		self.merged_cells.add(range_string)
    # Apply monkey patch
    worksheet.worksheet.merge_cells = merge_cells
patch_worksheet()


def date_and_time_from_timeStamp(cursor, time_stamp):
     # date_value = time_stamp[:time_stamp.find(' ')]
     # time_value = time_stamp[time_stamp.find(' ')+1:]
     # date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
     cursor.execute("SELECT DateFormat FROM " + 'tdstudycaseinfo')
     time_format = cursor.fetchall()
     if time_format[0][0] == 0:
         date_value = time_stamp[:time_stamp.find(' ')]
         time_value = time_stamp[time_stamp.find(' ') + 1:]
         date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
     elif time_format[0][0] == 1:
         date_value = time_stamp[:time_stamp.find(' ')]
         time_value = time_stamp[time_stamp.find(' ') + 1:]
         date_final = datetime.strptime(date_value, '%d-%m-%Y').date()
     elif time_format[0][0] == 2:
         date_value = time_stamp[:time_stamp.find(' ')]
         time_value = time_stamp[time_stamp.find(' ') + 1:]
         date_final = datetime.strptime(date_value, '%Y-%m-%d').date()
     else:
         date_value = time_stamp[:11]
         time_value = time_stamp[12:]
         date_final = datetime.strptime(date_value, '%m %d, %Y').date()
     return date_final, time_value


def is_tolerance_applied(db_value):
    if db_value == 0:
        return "No"
    else:
        return "Yes"

def apply_global_or_indivudual_tolerance(db_value):
    if db_value == 1:
        return "Global"
    else:
        return "Individual"


def get_time_stamp_from_result_id(table_names_list, cursor):
    table_index = table_names_list.index('tdtimeid')
    command = "SELECT * FROM " + table_names_list[table_index]
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()

    result_id_to_timestamp_dict = dict()
    result_id_list = []
    timestamp_list = []

    for i in range(len(sql_all_rows_data)):
        result_id_list.append(sql_all_rows_data[i][3])
        timestamp_list.append(sql_all_rows_data[i][2])
    
    timestamp_resultID_list = zip(result_id_list, timestamp_list)
    resultID_to_timestamp_dict = collections.defaultdict(list)
    for v, k in timestamp_resultID_list:
        resultID_to_timestamp_dict[v].append(k)
    return resultID_to_timestamp_dict

# IR-54886 - TDULF, report, date format should be consistent with project setting - START        
def get_date_format(table_names_list, cursor):
    Column_sc_info_date_format = 29

    # Getting all data from study case table
    sc_info_table_index = table_names_list.index('tdstudycaseinfo')
    sc_info_command = "SELECT * FROM " + table_names_list[sc_info_table_index]
    cursor.execute(sc_info_command)
    sc_info_sql_all_rows_data = cursor.fetchall()
    sc_info_total_rows_in_db = len(sc_info_sql_all_rows_data)

    date_sc_format = sc_info_sql_all_rows_data[0][Column_sc_info_date_format]
    return date_sc_format


def get_date_in_project_format(date_raw_format, date_sc_format):

    # Date format:
    #     0 --> mm-dd-yyyy
    #     1 --> dd-mm-yyyy
    #     2 --> yyyy-mm-dd
    #     3 --> mm dd, yyyy

    month_value = date_raw_format.split('-')[0]
    date_value = date_raw_format.split('-')[1]
    year_value = date_raw_format.split('-')[2] 

    date_format_to_value_dict = {
                                    0 : month_value + '-' + date_value + '-' + year_value,
                                    1 : date_value + '-' + month_value + '-' + year_value,
                                    2 : year_value + '-' + month_value + '-' + date_value,
                                    3 : month_value + ' ' + date_value  + ',' + year_value 
                                }
    

    return date_format_to_value_dict[date_sc_format]
# IR-54886 - TDULF, report, date format should be consistent with project setting - END

def check_db_version(table_names_list, cursor):
    Column_sc_info_date_format = 1

    # Getting all data from study case table
    sc_info_table_index = table_names_list.index('dbversion')
    sc_info_command = "SELECT * FROM " + table_names_list[sc_info_table_index]
    cursor.execute(sc_info_command)
    sc_info_sql_all_rows_data = cursor.fetchall()
    current_db_version = sc_info_sql_all_rows_data[0][1]
    coded_db_version = 100053.0
    if(current_db_version == coded_db_version):
        pass
    else:
        raise Exception("This Script is not compatible with the report db")

def fetch_all_data(cursor, table):
    command = "SELECT * FROM " + table
    cursor.execute(command)
    return cursor.fetchall()


def fetch_selected_columns(cursor, table, columns):
    for i in range(len(columns)):
        columns[i] = "[" + str(columns[i]) + "]"
    column = ','.join(columns)
    command = "SELECT " + column + " FROM " + table
    cursor.execute(command)
    return cursor.fetchall()


def fetch_conditional_columns(cursor, table, columns, condition):
    for i in range(len(columns)):
        columns[i] = "[" + str(columns[i]) + "]"
    column = ','.join(columns)
    command = "SELECT " + column + " FROM " + table + " Where " + condition
    cursor.execute(command)
    return cursor.fetchall()


def fetch_one_column(cursor, table, column):
    column = "[" + str(column) + "]"
    command = "SELECT " + column + " FROM " + table
    cursor.execute(command)
    return cursor.fetchall()


def get_updated_coordinates_inc(coordinates, inc):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]  # get second dollar sign index
    int_part = int(coordinates[second_emp_index + 1:]) + inc  # get the item after dollar sign, + 1 means next row
    # invalid literal for int() with base 10: '3:$G$3'
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates


def get_updated_columns_coordinates(coordinates, result_init_row, inc):
    short_coordinates = coordinates.replace('$', '')
    col = column_index_from_string((coordinate_from_string(short_coordinates))[0])
    col_inx = int(col) + inc
    col = _get_column_letter(col_inx)
    return absolute_coordinate(col + str(result_init_row))


def get_column_letter(coordinates):
    short_coordinates = coordinates.replace('$', '')
    col = column_index_from_string((coordinate_from_string(short_coordinates))[0])
    col_inx = int(col)
    col = _get_column_letter(col_inx)
    return col


def name_to_coordinate(wb):
    coordinates_dict = dict()
    for dn in wb.defined_names.definedName:  # create dictionary for defined names
        sheet_name, coordinate = dn.attr_text.split("!")
        sheet_name = sheet_name.replace("'", "")  # Avoid extra "'" generated when slicing string
        coordinate = get_updated_coordinates_inc(coordinate, 1)
        coordinates_dict[dn.name] = [sheet_name, coordinate]
    return coordinates_dict


def get_row_style(wb, ws, coordinates_dict):
    row_style = dict()
    for keys in coordinates_dict.keys():
        dn_sheet_name = coordinates_dict[keys][0]
        if dn_sheet_name == ws._WorkbookChild__title:
            cell = ws[coordinates_dict[keys][1]]
            row_style[keys] = cell._style
    return row_style


def read_time_format(cursor, time, option=0):
    # option 0, date + time default
    # option 1, date only
    cursor.execute("SELECT DateFormat FROM " + 'tdstudycaseinfo')
    time_format = cursor.fetchall()
    if option == 0:
        if time_format[0][0] == 0:
            return datetime.strptime(time, '%m-%d-%Y %H:%M:%S.%f')
        elif time_format[0][0] == 1:
            return datetime.strptime(time, '%d-%m-%Y %H:%M:%S.%f')
        elif time_format[0][0] == 2:
            return datetime.strptime(time, '%Y-%m-%d %H:%M:%S.%f')
        else:
            return datetime.strptime(time, '%m %d, %Y %H:%M:%S.%f')
    elif option == 1:
        if time_format[0][0] == 0:
            return datetime.strptime(time, '%m-%d-%Y')
        elif time_format[0][0] == 1:
            return datetime.strptime(time, '%d-%m-%Y')
        elif time_format[0][0] == 2:
            return datetime.strptime(time, '%Y-%m-%d')
        else:
            return datetime.strptime(time, '%m %d, %Y')


def is_milli_sec(cursor):
    cursor.execute("SELECT SimulationStepUnit, MillisecScale FROM " + 'tdstudycaseinfo')
    result = cursor.fetchall()
    time_step, millisec = result[0][0], result[0][1]
    if time_step == 'MilliSecond' and millisec == 1:
        return True
    else:
        return False
