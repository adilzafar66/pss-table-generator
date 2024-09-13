#***********************
#
# Copyright (c) 2021-2023, Operation Technology, Inc.
#
# THIS PROGRAM IS CONFIDENTIAL AND PROPRIETARY TO OPERATION TECHNOLOGY, INC. 
# ANY USE OF THIS PROGRAM IS SUBJECT TO THE PROGRAM SOFTWARE LICENSE AGREEMENT, 
# EXCEPT THAT THE USER MAY MODIFY THE PROGRAM FOR ITS OWN USE. 
# HOWEVER, THE PROGRAM MAY NOT BE REPRODUCED, PUBLISHED, OR DISCLOSED TO OTHERS 
# WITHOUT THE PRIOR WRITTEN CONSENT OF OPERATION TECHNOLOGY, INC.
#
#***********************

import sqlite3 
import os
import sys
import logging
from datetime import datetime 
from openpyxl import workbook, worksheet
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, _get_column_letter, \
    absolute_coordinate

def get_command_line_args(standard):
    this_py_file = sys.argv[0]
    report_db = sys.argv[1]
    final_sheet = True if sys.argv[2] == 'True' else False

    report_db_no_extension = os.path.splitext(report_db)[0]
    final_excel_file = report_db_no_extension + '_Report.xlsx'
    this_py_path = os.path.dirname(os.path.abspath(this_py_file))
    excel_template_file = this_py_path + '\\' + standard + ' DC Short Circuit Report_Template.xlsx'

    return this_py_file, report_db, final_excel_file, excel_template_file, final_sheet

def log_error(final_excel_file, e):
    file_name_w_ext = os.path.abspath(final_excel_file)
    file_name, extension = os.path.splitext(file_name_w_ext)
    log_file_name = file_name+'.log'
    logging.basicConfig(filename = log_file_name,filemode ='w', level = logging.ERROR, format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
    logger = logging.getLogger(__name__)   
    logger.exception(e.args[0])
    os.system(r'start notepad.exe ' + '"' + log_file_name +'"')

def deselect_all_cells(ws):
    ws.sheet_view.selection[0].activeCell = 'ZZ1'
    ws.sheet_view.selection[0].sqref = 'ZZ1'

def merge_cells(ws, coordinate1, coordinate2):
    coord1 = coordinate1.replace('$', '')
    coord2 = coordinate2.replace('$', '')
    ws.merge_cells(coord1 + ':' + coord2)

def increment_column_by_one(coordinate):
     first_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][0]
     second_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][1]
     column_alphabet = coordinate[first_emp_index + 1:second_emp_index]
     if len(column_alphabet) == 1:
         if column_alphabet == 'Z':
             start_writing_column = 'AA'
         else:
             start_writing_column = chr(ord(column_alphabet)+1)

     elif len(column_alphabet) == 2:
         if column_alphabet == 'ZZ':
             start_writing_column = 'AAA'
         else:
             first_char_of_column_alphabet = column_alphabet[:len(column_alphabet)-1]
             second_char_of_column_alphabet = column_alphabet[len(column_alphabet)-1:]
             start_writing_column = first_char_of_column_alphabet + chr(ord(second_char_of_column_alphabet)+1)

     start_writing_coordinate = coordinate[:first_emp_index+1] + start_writing_column + coordinate[second_emp_index:]
     return start_writing_coordinate

def decrement_column_by_one(coordinate):
     first_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][0]
     second_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][1]
     column_alphabet = coordinate[first_emp_index + 1:second_emp_index]
     start_writing_column = chr(ord(column_alphabet) - 1)
     start_writing_coordinate = coordinate[:first_emp_index + 1] + start_writing_column + coordinate[second_emp_index:]
     return start_writing_coordinate

def increment_row_by_one(coordinate):
    coord = increment_coordinates_by_number_of_rows(coordinate, 1)
    return coord

def decrement_row_by_one(coordinate):
    coord = decrement_coordinates_by_number_of_rows(coordinate, 1)
    return coord

def increment_coordinates_by_number_of_rows(coordinates, number_of_rows):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) + number_of_rows
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates

def decrement_coordinates_by_number_of_rows(coordinates, number_of_rows):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) - number_of_rows
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates

def get_excelName_to_coordinates_dict(wb, db_column_name_list):
    column_names = len(wb.defined_names.definedName)    
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(wb.defined_names.definedName[i].name)           
    
    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]
    
    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix+db_column_name_list[count]] = db_column_name_list[count]
    
    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            excel_column_name = defined_names_for_other_sheet[i] #wb.defined_names.definedName[i].name
            actual_coordinates = next(wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+2
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue
    return excel_column_name_to_coordinates

def connect_to_project(project_path):
    # Connecting to SQLite DB
    project = sqlite3.connect(project_path)
    cursor = project.cursor()

    # Get all the table names from report Database
    all_tables_name = project.execute("SELECT name FROM sqlite_master WHERE type = 'table';")
    table_names_list = []
    for table_names in all_tables_name:
        table_names_list.append(table_names[0])

    return cursor, table_names_list

def patch_worksheet():
    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            raise ValueError()
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
    									start_row,
    									get_column_letter(end_column),
    									end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError()
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self.merged_cells:
            self.merged_cells.add(range_string)

    # Apply monkey patch
    worksheet.worksheet.merge_cells = merge_cells
    
patch_worksheet()

def check_db_version(table_names_list, cursor):
    Column_sc_info_date_format = 1

    # Getting all data from study case table
    sc_info_table_index = table_names_list.index('dbversion')
    sc_info_command = "SELECT * FROM " + table_names_list[sc_info_table_index]
    cursor.execute(sc_info_command)
    sc_info_sql_all_rows_data = cursor.fetchall()
    current_db_version = sc_info_sql_all_rows_data[0][1]
    coded_db_version = 100053.0
    if(current_db_version == coded_db_version):
        pass
    else:
        raise Exception("This Script is not compatible with the report db")

def fetch_all_data(cursor, table):
    command = "SELECT * FROM " + table
    cursor.execute(command)
    return cursor.fetchall()

def fetch_selected_columns(cursor, table, columns):
    for i in range(len(columns)):
        columns[i] = "[" + str(columns[i]) + "]"
    column = ','.join(columns)
    command = "SELECT " + column + " FROM " + table
    cursor.execute(command)
    return cursor.fetchall()

def fetch_conditional_columns(cursor, table, columns, condition):
    for i in range(len(columns)):
        columns[i] = "[" + str(columns[i]) + "]"
    column = ','.join(columns)
    command = "SELECT " + column + " FROM " + table + " Where " + condition
    cursor.execute(command)
    return cursor.fetchall()

def fetch_one_column(cursor, table, column):
    column = "[" + str(column) + "]"
    command = "SELECT " + column + " FROM " + table
    cursor.execute(command)
    return cursor.fetchall()

def get_column_letter(coordinates):
    short_coordinates = coordinates.replace('$', '')
    col = column_index_from_string((coordinate_from_string(short_coordinates))[0])
    col_inx = int(col)
    col = _get_column_letter(col_inx)
    return col

def get_updated_coordinates_inc(coordinates, inc):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]  # get second dollar sign index
    int_part = int(coordinates[second_emp_index + 1:]) + inc  # get the item after dollar sign, + 1 means next row
    # invalid literal for int() with base 10: '3:$G$3'
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates    

def name_to_coordinate(wb):
    coordinates_dict = dict()
    for dn in wb.defined_names.definedName:  # create dictionary for defined names
        sheet_name, coordinate = dn.attr_text.split("!")
        sheet_name = sheet_name.replace("'", "")  # Avoid extra "'" generated when slicing string
        coordinate = get_updated_coordinates_inc(coordinate, 1)
        coordinates_dict[dn.name] = [sheet_name, coordinate]
    return coordinates_dict

def get_row_style(wb, ws, coordinates_dict):
    row_style = dict()
    for keys in coordinates_dict.keys():
        dn_sheet_name = coordinates_dict[keys][0]
        if dn_sheet_name == ws._WorkbookChild__title:
            cell = ws[coordinates_dict[keys][1]]
            row_style[keys] = cell._style
    return row_style
