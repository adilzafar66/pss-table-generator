#***********************
#
# Copyright (c) 2021-2023, Operation Technology, Inc.
#
# THIS PROGRAM IS CONFIDENTIAL AND PROPRIETARY TO OPERATION TECHNOLOGY, INC. 
# ANY USE OF THIS PROGRAM IS SUBJECT TO THE PROGRAM SOFTWARE LICENSE AGREEMENT, 
# EXCEPT THAT THE USER MAY MODIFY THE PROGRAM FOR ITS OWN USE. 
# HOWEVER, THE PROGRAM MAY NOT BE REPRODUCED, PUBLISHED, OR DISCLOSED TO OTHERS 
# WITHOUT THE PRIOR WRITTEN CONSENT OF OPERATION TECHNOLOGY, INC.
#
#***********************


import sqlite3
from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl import workbook
from openpyxl.utils import range_boundaries
import os
from datetime import datetime

def LoadProjectInfoData(wb, ws, cursor, table_names_list, templatePath):
    # Getting train study case info table
    table_index = table_names_list.index('trainstudycaseinfo')
    command = "SELECT * FROM " + table_names_list[table_index]
    cursor.execute(command)
    study_case_info_data = cursor.fetchall()
    total_rows_study_case_info = len(study_case_info_data)
    study_case_info_column_name_list = [description[0] for description in cursor.description]

    # Getting train header info table
    table_index = table_names_list.index('trainheadrinfo')
    command = "SELECT * FROM " + table_names_list[table_index]
    cursor.execute(command)
    header_info_data = cursor.fetchall()
    total_rows_header_info = len(header_info_data)
    header_info_column_name_list = [description[0] for description in cursor.description]

    column_names = len(wb.defined_names.definedName)
    defined_names_for_project_info = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] == "ProjectInfo":
            defined_names_for_project_info.append(wb.defined_names.definedName[i].name) 

    # nameActive = openpyxl.writer.workbook.get_active_sheet(wb)
    
    excel_column_name_to_coordinates_info = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, len(defined_names_for_project_info)):
        try:
            info_excel_column_name = defined_names_for_project_info[i]
            info_actual_coordinates = next(wb.defined_names[info_excel_column_name].destinations)[1]
            info_start_writing_coordinate = increase_column_by_one(info_actual_coordinates)
            excel_column_name_to_coordinates_info[info_excel_column_name] = info_start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    # Writing data for all the fields from header table in excel    
    coordinates_for_projectName = excel_column_name_to_coordinates_info["ProjectInfo_Project"]
    ws[coordinates_for_projectName] = header_info_data[0][0]

    coordinates_for_PSRev = excel_column_name_to_coordinates_info["ProjectInfo_PSRev"]
    ws[coordinates_for_PSRev] = header_info_data[0][4]

    coordinates_for_Loc = excel_column_name_to_coordinates_info["ProjectInfo_Loc"]
    ws[coordinates_for_Loc] = header_info_data[0][1]

    coordinates_for_Contract = excel_column_name_to_coordinates_info["ProjectInfo_Contr"]
    ws[coordinates_for_Contract] = header_info_data[0][2]

    coordinates_for_Date = excel_column_name_to_coordinates_info["ProjectInfo_Date"]
    ws[coordinates_for_Date] = header_info_data[0][6]

    # Writing data for all the fields from StudyCase Info table in excel    
    coordinates_for_fileName = excel_column_name_to_coordinates_info["ProjectInfo_FileN"]
    ws[coordinates_for_fileName] = study_case_info_data[0][4]

    coordinates_for_filePath = excel_column_name_to_coordinates_info["ProjectInfo_FilePath"]
    ws[coordinates_for_filePath] = study_case_info_data[0][5]

    coordinates_for_libFileName = excel_column_name_to_coordinates_info["ProjectInfo_LibFileN"]
    ws[coordinates_for_libFileName] = study_case_info_data[0][8]

    coordinates_for_libFilePath = excel_column_name_to_coordinates_info["ProjectInfo_LibFilePath"]
    ws[coordinates_for_libFilePath] = study_case_info_data[0][9]

    coordinates_for_warehouseFileName = excel_column_name_to_coordinates_info["ProjectInfo_WarehouseFileN"]
    ws[coordinates_for_warehouseFileName] = study_case_info_data[0][10]

    coordinates_for_warehouseFilePath = excel_column_name_to_coordinates_info["ProjectInfo_WarehouseFilePath"]
    ws[coordinates_for_warehouseFilePath] = study_case_info_data[0][11]

    coordinates_for_Output = excel_column_name_to_coordinates_info["ProjectInfo_Output"]
    ws[coordinates_for_Output] = study_case_info_data[0][6]

    coordinates_for_OutputFilePath = excel_column_name_to_coordinates_info["ProjectInfo_OutputFilePath"]
    ws[coordinates_for_OutputFilePath] = study_case_info_data[0][7]

    coordinates_for_config = excel_column_name_to_coordinates_info["ProjectInfo_Config"]
    ws[coordinates_for_config] = study_case_info_data[0][12]

    coordinates_for_revision = excel_column_name_to_coordinates_info["ProjectInfo_Revision"]
    ws[coordinates_for_revision] = study_case_info_data[0][13]

    coordinates_for_studyCaseID = excel_column_name_to_coordinates_info["ProjectInfo_StudyCaseID"]
    ws[coordinates_for_studyCaseID] = study_case_info_data[0][0]

    coordinates_for_standard = excel_column_name_to_coordinates_info["ProjectInfo_Standard"]
    ws[coordinates_for_standard] = study_case_info_data[0][3]

    coordinates_for_freq = excel_column_name_to_coordinates_info["ProjectInfo_Freq"]
    ws[coordinates_for_freq] = study_case_info_data[0][1]

    path_name = os.path.dirname(templatePath)
    path_name2 = os.path.dirname(path_name)
    imageLocation = path_name2 + r'\eTrax_logo.jpg'

    img = openpyxl.drawing.image.Image(imageLocation)
    ws.add_image(img, 'G3')
    ws.sheet_view.showGridLines = False
    
    patch_worksheet()  

    
def increase_column_by_one(coordinate):
     first_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][0]
     second_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][1]
     column_alphabet = coordinate[first_emp_index + 1:second_emp_index]
     if len(column_alphabet) == 1:
         if column_alphabet == 'Z':
             start_writing_column = 'AA'
         else:
             start_writing_column = chr(ord(column_alphabet)+1)

     elif len(column_alphabet) == 2:
         if column_alphabet == 'ZZ':
             start_writing_column = 'AAA'
         else:
             first_char_of_column_alphabet = column_alphabet[:len(column_alphabet)-1]
             second_char_of_column_alphabet = column_alphabet[len(column_alphabet)-1:]
             start_writing_column = first_char_of_column_alphabet + chr(ord(second_char_of_column_alphabet)+1)

     start_writing_coordinate = coordinate[:first_emp_index+1] + start_writing_column + coordinate[second_emp_index:]
     return start_writing_coordinate


def get_excelName_to_coordinates_dict(wb, db_column_name_list):
    column_names = len(wb.defined_names.definedName)    
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(wb.defined_names.definedName[i].name)           
    
    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]
    
    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix+db_column_name_list[count]] = db_column_name_list[count]
    
    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            excel_column_name = defined_names_for_other_sheet[i] #wb.defined_names.definedName[i].name
            actual_coordinates = next(wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+2
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue
    return excel_column_name_to_coordinates


def connect_to_project(project_path):
    # ....Connecting to SQLite DB....
    project = sqlite3.connect(project_path)
    cursor = project.cursor()

    # ....Get all the table names from report Database....
    all_tables_name = project.execute("SELECT name FROM sqlite_master WHERE type = 'table';")
    table_names_list = []
    for table_names in all_tables_name:
        table_names_list.append(table_names[0])

    return cursor, table_names_list


def get_updated_coordinates(coordinates):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) + 1
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates


def get_updated_coordinates_by_value(coordinates, value):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) + value
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates


def get_actual_day(int_val):
    int_to_day = {0: "Mon.", 1: "Tue.", 2: "Wed.", 3: "Thurs.", 4: "Fri.", 5: "Sat.", 6: "Sun."}
    value_day = int_to_day[int_val]
    return value_day


def get_phase_value(raw_phase):    
    raw_phase_to_actualphase = {
        # bus/branch/load connection types
        0: "ABC",
        3: "DC" ,
        5: "A"  ,
        6: "B"  ,
        7: "C"  ,
        8: "AB" ,
        9: "BC" ,
        10: "CA" ,
        
        # bus and branch under the center-tap Xfmr
        11: "A3W" ,
        12: "B3W" ,
        13: "C3W" ,
        14: "AB3W",
        15: "BC3W",
        16: "CA3W",
        20: "LL3W",
        
        # all elements: AB->LL,BC->L2,CA->L1 to handle LL2W
        17: "LL"       ,
        18: "L1"       ,
        19: "L2"       ,
        22: "AT2W"     ,
        23: "BT2W"     ,
        24: "CT2W"     ,
        25: "ABT2W"    ,
        26: "BCT2W"    ,
        27: "CAT2W"    ,
        28: "AT2WN"    ,
        29: "BT2WN"    ,
        30: "CT2WN"    ,
        31: "ABT2WN"   ,
        32: "BCT2WN"   ,
        33: "CAT2WN"   ,
        34: "AT3WB"    ,
        35: "BT3WB"    ,
        36: "CT3WB"    ,
        37: "ABT3WB"   ,
        38: "BCT3WB"   ,
        39: "CAT3WB"   ,
        40: "AT3WN"    ,
        41: "BT3WN"    ,
        42: "CT3WN"    ,
        43: "ABT3WN"   ,
        44: "BCT3WN"   ,
        45: "CAT3WN"   ,
        46: "ScottT2W" ,
        47: "ScottT2WN",
        48: "ScottT3WB",
        49: "ScottT3WN",
        }
    values = raw_phase_to_actualphase[raw_phase]
    return values


def patch_worksheet():
	def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
		if not range_string and not all((start_row, start_column, end_row, end_column)):
			raise ValueError()
		elif not range_string:
			range_string = '%s%s:%s%s' % (get_column_letter(start_column),
										start_row,
										get_column_letter(end_column),
										end_row)
		elif ":" not in range_string:
			if COORD_RE.match(range_string):
				return  # Single cell, do nothing
			raise ValueError()
		else:
			range_string = range_string.replace('$', '')
	
		if range_string not in self.merged_cells:
			self.merged_cells.add(range_string)
	# Apply monkey patch
	worksheet.worksheet.merge_cells = merge_cells
patch_worksheet()

def date_and_time_from_timeStamp(time_stamp):
     # date_value = time_stamp[:time_stamp.find(' ')]
     time_value = time_stamp[time_stamp.find(' ')+1:]
     # date_final = datetime.strptime(date_value, '%Y-%m-%d').date()
     return time_value
