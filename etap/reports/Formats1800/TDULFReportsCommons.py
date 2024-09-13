#***********************
#
# Copyright (c) 2021-2023, Operation Technology, Inc.
#
# THIS PROGRAM IS CONFIDENTIAL AND PROPRIETARY TO OPERATION TECHNOLOGY, INC. 
# ANY USE OF THIS PROGRAM IS SUBJECT TO THE PROGRAM SOFTWARE LICENSE AGREEMENT, 
# EXCEPT THAT THE USER MAY MODIFY THE PROGRAM FOR ITS OWN USE. 
# HOWEVER, THE PROGRAM MAY NOT BE REPRODUCED, PUBLISHED, OR DISCLOSED TO OTHERS 
# WITHOUT THE PRIOR WRITTEN CONSENT OF OPERATION TECHNOLOGY, INC.
#
#***********************


import sqlite3 
from datetime import datetime 
from openpyxl import workbook, worksheet

def increase_column_by_one(coordinate):
     first_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][0]
     second_emp_index = [i for i, ltr in enumerate(coordinate) if ltr == '$'][1]
     column_alphabet = coordinate[first_emp_index + 1:second_emp_index]
     if len(column_alphabet) == 1:
         if column_alphabet == 'Z':
             start_writing_column = 'AA'
         else:
             start_writing_column = chr(ord(column_alphabet)+1)

     elif len(column_alphabet) == 2:
         if column_alphabet == 'ZZ':
             start_writing_column = 'AAA'
         else:
             first_char_of_column_alphabet = column_alphabet[:len(column_alphabet)-1]
             second_char_of_column_alphabet = column_alphabet[len(column_alphabet)-1:]
             start_writing_column = first_char_of_column_alphabet + chr(ord(second_char_of_column_alphabet)+1)

     start_writing_coordinate = coordinate[:first_emp_index+1] + start_writing_column + coordinate[second_emp_index:]
     return start_writing_coordinate


def get_excelName_to_coordinates_dict(wb, db_column_name_list):
    column_names = len(wb.defined_names.definedName)    
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(wb.defined_names.definedName[i].name)           
    
    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]
    
    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix+db_column_name_list[count]] = db_column_name_list[count]
    
    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            excel_column_name = defined_names_for_other_sheet[i] #wb.defined_names.definedName[i].name
            actual_coordinates = next(wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+2
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue
    return excel_column_name_to_coordinates


def connect_to_project(project_path):
    # ....Connecting to SQLite DB....
    project = sqlite3.connect(project_path)
    cursor = project.cursor()

    # ....Get all the table names from report Database....
    all_tables_name = project.execute("SELECT name FROM sqlite_master WHERE type = 'table';")
    table_names_list = []
    for table_names in all_tables_name:
        table_names_list.append(table_names[0])

    return cursor, table_names_list


def get_updated_coordinates(coordinates):
    second_emp_index = [i for i, ltr in enumerate(coordinates) if ltr == '$'][1]
    int_part = int(coordinates[second_emp_index + 1:]) + 1
    coordinates = coordinates[:second_emp_index + 1] + str(int_part)
    return coordinates


def get_actual_day(int_val):
    int_to_day = {0: "Mon.", 1: "Tue.", 2: "Wed.", 3: "Thurs.", 4: "Fri.", 5: "Sat.", 6: "Sun."}
    value_day = int_to_day[int_val]
    return value_day


def patch_worksheet():
    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
    	if not range_string and not all((start_row, start_column, end_row, end_column)):
    		raise ValueError()
    	elif not range_string:
    		range_string = '%s%s:%s%s' % (get_column_letter(start_column),
    									start_row,
    									get_column_letter(end_column),
    									end_row)
    	elif ":" not in range_string:
    		if COORD_RE.match(range_string):
    			return  # Single cell, do nothing
    		raise ValueError()
    	else:
    		range_string = range_string.replace('$', '')
    
    	if range_string not in self.merged_cells:
    		self.merged_cells.add(range_string)
    # Apply monkey patch
    worksheet.worksheet.merge_cells = merge_cells
patch_worksheet()


def date_and_time_from_timeStamp(time_stamp):
     date_value = time_stamp[:time_stamp.find(' ')]
     time_value = time_stamp[time_stamp.find(' ')+1:]
     date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
     return date_final, time_value


def is_tolerance_applied(db_value):
    if db_value == 0:
        return "No"
    else:
        return "Yes"

def apply_global_or_indivudual_tolerance(db_value):
    if db_value == 1:
        return "Global"
    else:
        return "Individual"


def get_time_stamp_from_result_id(table_names_list, cursor):
    table_index = table_names_list.index('tdtimeid')
    command = "SELECT * FROM " + table_names_list[table_index]
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()

    result_id_to_timestamp_dict = dict()
    result_id_list = []
    timestamp_list = []

    for i in range(len(sql_all_rows_data)):
        result_id_list.append(sql_all_rows_data[i][3])
        timestamp_list.append(sql_all_rows_data[i][2])
    
    timestamp_resultID_list = zip(result_id_list, timestamp_list)
    resultID_to_timestamp_dict = collections.defaultdict(list)
    for v, k in timestamp_resultID_list:
        resultID_to_timestamp_dict[v].append(k)
    return resultID_to_timestamp_dict
