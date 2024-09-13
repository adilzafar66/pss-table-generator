#***********************
#
# Copyright (c) 2021-2023, Operation Technology, Inc.
#
# THIS PROGRAM IS CONFIDENTIAL AND PROPRIETARY TO OPERATION TECHNOLOGY, INC. 
# ANY USE OF THIS PROGRAM IS SUBJECT TO THE PROGRAM SOFTWARE LICENSE AGREEMENT, 
# EXCEPT THAT THE USER MAY MODIFY THE PROGRAM FOR ITS OWN USE. 
# HOWEVER, THE PROGRAM MAY NOT BE REPRODUCED, PUBLISHED, OR DISCLOSED TO OTHERS 
# WITHOUT THE PRIOR WRITTEN CONSENT OF OPERATION TECHNOLOGY, INC.
#
#***********************

from openpyxl import load_workbook, worksheet
from shutil import copyfile
import time
from etap.scripting.reports import TDULFReportsCommons as comms
import sys
import os
from datetime import datetime
from copy import copy
import collections
import openpyxl


def load_events_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    # SQLite DB Column definations
    column_event_id = 1  # Entry 2 of sql_all_rows_data.
    column_time = 2
    column_device_type = 3
    column_device_id = 4
    column_action = 5
    column_action_percent = 7
    column_action_time = 8

    # Getting all data from events table
    # table_index = table_names_list.index('TDActions')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDActions"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)

    db_column_name_list = [description[0]
                           for description in cursor.description]

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    comms.patch_worksheet()

    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix +
                                      db_column_name_list[count]] = db_column_name_list[count]

    excel_column_name_to_coordinates = comms.get_excelName_to_coordinates_dict(
        wb, db_column_name_list)

    coordinates_for_EventID = excel_column_name_to_coordinates["TDEvents_EventID"]
    coordinates_for_Time = excel_column_name_to_coordinates["TDEvents_Time"]
    coordinates_for_Date = excel_column_name_to_coordinates["TDEvents_Date"]
    coordinates_for_DeviceType = excel_column_name_to_coordinates["TDEvents_DeviceType"]
    coordinates_for_DeviceID = excel_column_name_to_coordinates["TDEvents_DeviceID"]
    coordinates_for_Action = excel_column_name_to_coordinates["TDEvents_Action"]
    coordinates_for_ActionPercent = excel_column_name_to_coordinates["TDEvents_ActionPercent"]
    coordinates_for_ActionTime = excel_column_name_to_coordinates["TDEvents_ActionTime"]

    first_row_style = dict()
    first_row_style["TDEvents_EventID"] = ws[comms.get_updated_coordinates(
        coordinates_for_EventID)]._style
    first_row_style["TDEvents_Time"] = ws[comms.get_updated_coordinates(
        coordinates_for_Time)]._style
    first_row_style["TDEvents_Date"] = ws[comms.get_updated_coordinates(
        coordinates_for_Date)]._style
    first_row_style["TDEvents_DeviceType"] = ws[comms.get_updated_coordinates(
        coordinates_for_DeviceType)]._style
    first_row_style["TDEvents_DeviceID"] = ws[comms.get_updated_coordinates(
        coordinates_for_DeviceID)]._style
    first_row_style["TDEvents_Action"] = ws[comms.get_updated_coordinates(
        coordinates_for_Action)]._style
    first_row_style["TDEvents_ActionPercent"] = ws[comms.get_updated_coordinates(
        coordinates_for_ActionPercent)]._style
    first_row_style["TDEvents_ActionTime"] = ws[comms.get_updated_coordinates(
        coordinates_for_ActionTime)]._style

    for i in range(0, total_rows_in_db):
        ws[coordinates_for_EventID]._style = copy(
            first_row_style["TDEvents_EventID"])
        ws[coordinates_for_EventID] = sql_all_rows_data[i][column_event_id]
        coordinates_for_EventID = comms.get_updated_coordinates(
            coordinates_for_EventID)

        ws[coordinates_for_Time]._style = copy(
            first_row_style["TDEvents_Time"])
        ws[coordinates_for_Time] = sql_all_rows_data[i][column_time].split(None, 1)[
            1]
        coordinates_for_Time = comms.get_updated_coordinates(
            coordinates_for_Time)

        ws[coordinates_for_Date]._style = copy(
            first_row_style["TDEvents_Date"])
        ws[coordinates_for_Date] = sql_all_rows_data[i][column_time].split(None, 1)[
            0]
        coordinates_for_Date = comms.get_updated_coordinates(
            coordinates_for_Date)

        ws[coordinates_for_DeviceType]._style = copy(
            first_row_style["TDEvents_DeviceType"])
        ws[coordinates_for_DeviceType] = sql_all_rows_data[i][column_device_id]
        coordinates_for_DeviceType = comms.get_updated_coordinates(
            coordinates_for_DeviceType)

        ws[coordinates_for_DeviceID]._style = copy(
            first_row_style["TDEvents_DeviceID"])
        ws[coordinates_for_DeviceID] = sql_all_rows_data[i][column_device_id]
        coordinates_for_DeviceID = comms.get_updated_coordinates(
            coordinates_for_DeviceID)

        ws[coordinates_for_Action]._style = copy(
            first_row_style["TDEvents_Action"])
        ws[coordinates_for_Action] = sql_all_rows_data[i][column_action]
        coordinates_for_Action = comms.get_updated_coordinates(
            coordinates_for_Action)

        ws[coordinates_for_ActionPercent]._style = copy(
            first_row_style["TDEvents_ActionPercent"])
        ws[coordinates_for_ActionPercent] = sql_all_rows_data[i][column_action_percent]
        coordinates_for_ActionPercent = comms.get_updated_coordinates(
            coordinates_for_ActionPercent)

        ws[coordinates_for_ActionTime]._style = copy(
            first_row_style["TDEvents_ActionTime"])
        ws[coordinates_for_ActionTime] = sql_all_rows_data[i][column_action_time]
        coordinates_for_ActionTime = comms.get_updated_coordinates(
            coordinates_for_ActionTime)


def load_device_profile_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    # Getting all data from system results table
    # table_index = table_names_list.index('tddevice')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDDevice"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)

    db_column_name_list = [description[0]
                           for description in cursor.description]

    # Connecting to Excel and Fetching Defined Names

    column_names = len(wb.defined_names.definedName)

    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    excel_column_name_to_coordinates = comms.get_excelName_to_coordinates_dict(
        wb, db_column_name_list)

    coordinates_for_deviceID = excel_column_name_to_coordinates["DeviceProfile_DeviceName"]
    coordinates_for_deviceType = excel_column_name_to_coordinates["DeviceProfile_DeviceType"]
    coordinates_for_dataType = excel_column_name_to_coordinates["DeviceProfile_DataType"]
    coordinates_for_type1 = excel_column_name_to_coordinates["DeviceProfile_Type1"]
    coordinates_for_type2 = excel_column_name_to_coordinates["DeviceProfile_Type2"]
    coordinates_for_energizeDate = excel_column_name_to_coordinates["DeviceProfile_EnergizeDate"]
    coordinates_for_OutofServiceDate = excel_column_name_to_coordinates[
        "DeviceProfile_OutofServiceDate"]
    coordinates_for_GrowthCurve = excel_column_name_to_coordinates["DeviceProfile_GrowthCurve"]
    coordinates_for_DateGapOptions = excel_column_name_to_coordinates[
        "DeviceProfile_DateGapOptions"]

    # Storing style of each cell in the second row in a dictionary.
    second_row_style = dict()  # {"range_string":cell_style}
    second_row_style["DeviceProfile_DeviceName"] = ws[(
        coordinates_for_deviceID)]._style
    second_row_style["DeviceProfile_DeviceType"] = ws[(
        coordinates_for_deviceType)]._style
    second_row_style["DeviceProfile_DataType"] = ws[(
        coordinates_for_dataType)]._style
    second_row_style["DeviceProfile_Type1"] = ws[(
        coordinates_for_type1)]._style
    second_row_style["DeviceProfile_Type2"] = ws[(
        coordinates_for_type2)]._style
    second_row_style["DeviceProfile_EnergizeDate"] = ws[(
        coordinates_for_energizeDate)]._style
    second_row_style["DeviceProfile_OutofServiceDate"] = ws[(
        coordinates_for_OutofServiceDate)]._style
    second_row_style["DeviceProfile_GrowthCurve"] = ws[(
        coordinates_for_GrowthCurve)]._style
    second_row_style["DeviceProfile_DateGapOptions"] = ws[(
        coordinates_for_DateGapOptions)]._style

    for i in range(0, total_rows_in_db):

        ws[coordinates_for_deviceID]._style = copy(
            second_row_style["DeviceProfile_DeviceName"])
        ws[coordinates_for_deviceID] = sql_all_rows_data[i][2]
        coordinates_for_deviceID = comms.get_updated_coordinates(
            coordinates_for_deviceID)

        ws[coordinates_for_deviceType]._style = copy(
            second_row_style["DeviceProfile_DeviceType"])
        ws[coordinates_for_deviceType] = sql_all_rows_data[i][3]
        coordinates_for_deviceType = comms.get_updated_coordinates(
            coordinates_for_deviceType)

        ws[coordinates_for_dataType]._style = copy(
            second_row_style["DeviceProfile_DataType"])
        ws[coordinates_for_dataType] = sql_all_rows_data[i][4]
        coordinates_for_dataType = comms.get_updated_coordinates(
            coordinates_for_dataType)

        ws[coordinates_for_type1]._style = copy(
            second_row_style["DeviceProfile_Type1"])
        ws[coordinates_for_type1] = sql_all_rows_data[i][5]
        coordinates_for_type1 = comms.get_updated_coordinates(
            coordinates_for_type1)

        ws[coordinates_for_type2]._style = copy(
            second_row_style["DeviceProfile_Type2"])
        ws[coordinates_for_type2] = sql_all_rows_data[i][6]
        coordinates_for_type2 = comms.get_updated_coordinates(
            coordinates_for_type2)

        ws[coordinates_for_energizeDate]._style = copy(
            second_row_style["DeviceProfile_EnergizeDate"])
        energize_date_value = sql_all_rows_data[i][7]
        energize_date_final = datetime.strptime(
            energize_date_value, '%Y-%m-%d').date()
        ws[coordinates_for_energizeDate] = energize_date_final.strftime('%x')
        coordinates_for_energizeDate = comms.get_updated_coordinates(
            coordinates_for_energizeDate)

        ws[coordinates_for_OutofServiceDate]._style = copy(
            second_row_style["DeviceProfile_OutofServiceDate"])
        outOfService_date_value_in_sc_format = sql_all_rows_data[i][8]
        ws[coordinates_for_OutofServiceDate] = outOfService_date_value_in_sc_format
        coordinates_for_OutofServiceDate = comms.get_updated_coordinates(
            coordinates_for_OutofServiceDate)

        ws[coordinates_for_GrowthCurve]._style = copy(
            second_row_style["DeviceProfile_GrowthCurve"])
        ws[coordinates_for_GrowthCurve] = sql_all_rows_data[i][9]
        coordinates_for_GrowthCurve = comms.get_updated_coordinates(
            coordinates_for_GrowthCurve)

        dataGapOptions_raw = sql_all_rows_data[i][10]
        dataGapOptionsFinal = 'Something other than 0, 1 or 2'
        if dataGapOptions_raw == 0:
            dataGapOptionsFinal = ''
        elif dataGapOptions_raw == 1:
            dataGapOptionsFinal = 'Use Last Value'
        elif dataGapOptions_raw == 2:
            dataGapOptionsFinal = 'Loading Category'

        ws[coordinates_for_DateGapOptions]._style = copy(
            second_row_style["DeviceProfile_DateGapOptions"])
        ws[coordinates_for_DateGapOptions] = dataGapOptionsFinal
        coordinates_for_DateGapOptions = comms.get_updated_coordinates(
            coordinates_for_DateGapOptions)


def load_alert_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    # SQLite DB Column definations
    column_result_id = 1  # Entry 2 of sql_all_rows_data.
    column_device_id = 2
    column_device_type = 3
    column_rated = 6
    column_unit = 5
    column_deviationA = 10
    column_deviationB = 11
    column_deviationC = 12
    column_condition = 13
    column_alert_type = 14

    # Getting all data from system results table
    # table_index = table_names_list.index('tdalert')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDAlert"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)

    db_column_name_list = [description[0]
                           for description in cursor.description]
    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    comms.patch_worksheet()

    excel_column_name_to_coordinates = dict()
    column_names = len(wb.defined_names.definedName)
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(
                wb.defined_names.definedName[i].name)

    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]

    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix +
                                      db_column_name_list[count]] = db_column_name_list[count]

    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            # wb.defined_names.definedName[i].name
            excel_column_name = defined_names_for_other_sheet[i]
            actual_coordinates = next(
                wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(
                actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+1
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(
                int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    coordinates_for_time = excel_column_name_to_coordinates["Alerts_Time"]
    coordinates_for_date = excel_column_name_to_coordinates["Alerts_Date"]
    coordinates_for_deviceID = excel_column_name_to_coordinates["Alerts_DeviceID"]
    coordinates_for_deviceType = excel_column_name_to_coordinates["Alerts_DeviceType"]
    coordinates_for_rated = excel_column_name_to_coordinates["Alerts_Rated"]
    coordinates_for_unit = excel_column_name_to_coordinates["Alerts_Unit"]
    coordinates_for_deviationA = excel_column_name_to_coordinates["Alerts_DeviationA"]
    coordinates_for_deviationB = excel_column_name_to_coordinates["Alerts_DeviationB"]
    coordinates_for_deviationC = excel_column_name_to_coordinates["Alerts_DeviationC"]
    coordinates_for_condition = excel_column_name_to_coordinates["Alerts_Condition"]
    coordinates_for_alertType = excel_column_name_to_coordinates["Alerts_AlertType"]

    # Storing style of each cell in the second row in a dictionary.
    second_row_style = dict()  # {"range_string":cell_style}
    second_row_style["Alerts_Time"] = ws[(coordinates_for_time)]._style
    second_row_style["Alerts_Date"] = ws[(coordinates_for_date)]._style
    second_row_style["Alerts_DeviceID"] = ws[(coordinates_for_deviceID)]._style
    second_row_style["Alerts_DeviceType"] = ws[(
        coordinates_for_deviceType)]._style
    second_row_style["Alerts_Rated"] = ws[(coordinates_for_rated)]._style
    second_row_style["Alerts_Unit"] = ws[(coordinates_for_unit)]._style
    second_row_style["Alerts_DeviationA"] = ws[(
        coordinates_for_deviationA)]._style
    second_row_style["Alerts_DeviationB"] = ws[(
        coordinates_for_deviationB)]._style
    second_row_style["Alerts_DeviationC"] = ws[(
        coordinates_for_deviationC)]._style
    second_row_style["Alerts_Condition"] = ws[(
        coordinates_for_condition)]._style
    second_row_style["Alerts_AlertType"] = ws[(
        coordinates_for_alertType)]._style

    for i in range(0, total_rows_in_db):
        # Convert time and date to saperate columns...
        time_stamp = resultID_to_timestamp_dict[sql_all_rows_data[i]
                                                [column_result_id]][0]
        date_value = time_stamp[:time_stamp.find(' ')]
        time_value = time_stamp[time_stamp.find(' ')+1:]
        date_final = datetime.strptime(date_value, '%m-%d-%Y').date()

        ws[coordinates_for_time]._style = copy(second_row_style["Alerts_Time"])
        ws[coordinates_for_time] = time_value
        coordinates_for_time = comms.get_updated_coordinates(
            coordinates_for_time)

        ws[coordinates_for_date]._style = copy(second_row_style["Alerts_Date"])
        ws[coordinates_for_date] = date_value
        coordinates_for_date = comms.get_updated_coordinates(
            coordinates_for_date)

        ws[coordinates_for_deviceID]._style = copy(
            second_row_style["Alerts_DeviceID"])
        ws[coordinates_for_deviceID] = sql_all_rows_data[i][column_device_id]
        coordinates_for_deviceID = comms.get_updated_coordinates(
            coordinates_for_deviceID)

        ws[coordinates_for_deviceType]._style = copy(
            second_row_style["Alerts_DeviceType"])
        ws[coordinates_for_deviceType] = sql_all_rows_data[i][column_device_type]
        coordinates_for_deviceType = comms.get_updated_coordinates(
            coordinates_for_deviceType)

        ws[coordinates_for_rated]._style = copy(
            second_row_style["Alerts_Rated"])
        ws[coordinates_for_rated] = sql_all_rows_data[i][column_rated]
        coordinates_for_rated = comms.get_updated_coordinates(
            coordinates_for_rated)

        ws[coordinates_for_unit]._style = copy(second_row_style["Alerts_Unit"])
        ws[coordinates_for_unit] = sql_all_rows_data[i][column_unit]
        coordinates_for_unit = comms.get_updated_coordinates(
            coordinates_for_unit)

        ws[coordinates_for_deviationA]._style = copy(
            second_row_style["Alerts_DeviationA"])
        ws[coordinates_for_deviationA] = sql_all_rows_data[i][column_deviationA]
        coordinates_for_deviationA = comms.get_updated_coordinates(
            coordinates_for_deviationA)

        ws[coordinates_for_deviationB]._style = copy(
            second_row_style["Alerts_DeviationB"])
        ws[coordinates_for_deviationB] = sql_all_rows_data[i][column_deviationB]
        coordinates_for_deviationB = comms.get_updated_coordinates(
            coordinates_for_deviationB)

        ws[coordinates_for_deviationC]._style = copy(
            second_row_style["Alerts_DeviationC"])
        ws[coordinates_for_deviationC] = sql_all_rows_data[i][column_deviationC]
        coordinates_for_deviationC = comms.get_updated_coordinates(
            coordinates_for_deviationC)

        ws[coordinates_for_condition]._style = copy(
            second_row_style["Alerts_Condition"])
        ws[coordinates_for_condition] = sql_all_rows_data[i][column_condition]
        coordinates_for_condition = comms.get_updated_coordinates(
            coordinates_for_condition)

        ws[coordinates_for_alertType]._style = copy(
            second_row_style["Alerts_AlertType"])
        ws[coordinates_for_alertType] = sql_all_rows_data[i][column_alert_type]
        coordinates_for_alertType = comms.get_updated_coordinates(
            coordinates_for_alertType)


def load_loading_worst_case_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    # Getting all data from system results table
    # table_index = table_names_list.index('td2tworstoverloadcases')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TD2TWorstOverLoadCases"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)

    db_column_name_list = [description[0]
                           for description in cursor.description]

    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    column_names = len(wb.defined_names.definedName)
    defined_names_for_other_sheet = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] != "ProjectInfo":
            defined_names_for_other_sheet.append(
                wb.defined_names.definedName[i].name)

    index_of_underscore = defined_names_for_other_sheet[1].find("_")
    excel_column_name_prefix = defined_names_for_other_sheet[1][:index_of_underscore+1]

    excel_to_sql_column_name_dict = dict()
    for count in range(len(db_column_name_list)):
        excel_to_sql_column_name_dict[excel_column_name_prefix +
                                      db_column_name_list[count]] = db_column_name_list[count]

    # Get coordinates for the column names
    excel_column_name_to_coordinates = dict()
    for i in range(0, len(defined_names_for_other_sheet)):
        try:
            # wb.defined_names.definedName[i].name
            excel_column_name = defined_names_for_other_sheet[i]
            actual_coordinates = next(
                wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(
                actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+1
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(
                int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    coordinates_for_deviceID = excel_column_name_to_coordinates["LoadWorstCase_DeviceID"]
    coordinates_for_deviceType = excel_column_name_to_coordinates["LoadWorstCase_DeviceType"]
    coordinates_for_rated = excel_column_name_to_coordinates["LoadWorstCase_Capacity"]
    coordinates_for_unit = excel_column_name_to_coordinates["LoadWorstCase_Unit"]
    coordinates_for_loadingA = excel_column_name_to_coordinates[
        "LoadWorstCase_MaxBranchLoadingPhA"]
    coordinates_for_loadingB = excel_column_name_to_coordinates[
        "LoadWorstCase_MaxBranchLoadingPhB"]
    coordinates_for_loadingC = excel_column_name_to_coordinates[
        "LoadWorstCase_MaxBranchLoadingPhC"]
    coordinates_for_date = excel_column_name_to_coordinates["LoadWorstCase_Date"]
    coordinates_for_time = excel_column_name_to_coordinates["LoadWorstCase_Time"]

    # Storing style of each cell in the second row in a dictionary.
    second_row_style = dict()  # {"range_string":cell_style}

    second_row_style["LoadWorstCase_DeviceID"] = ws[comms.get_updated_coordinates(
        coordinates_for_deviceID)]._style
    second_row_style["LoadWorstCase_DeviceType"] = ws[comms.get_updated_coordinates(
        coordinates_for_deviceType)]._style
    second_row_style["LoadWorstCase_Capacity"] = ws[comms.get_updated_coordinates(
        coordinates_for_rated)]._style
    second_row_style["LoadWorstCase_Unit"] = ws[comms.get_updated_coordinates(
        coordinates_for_unit)]._style
    second_row_style["LoadWorstCase_MaxBranchLoadingPhA"] = ws[comms.get_updated_coordinates(
        coordinates_for_loadingA)]._style
    second_row_style["LoadWorstCase_MaxBranchLoadingPhB"] = ws[comms.get_updated_coordinates(
        coordinates_for_loadingB)]._style
    second_row_style["LoadWorstCase_MaxBranchLoadingPhC"] = ws[comms.get_updated_coordinates(
        coordinates_for_loadingC)]._style
    second_row_style["LoadWorstCase_Date"] = ws[comms.get_updated_coordinates(
        coordinates_for_date)]._style
    second_row_style["LoadWorstCase_Time"] = ws[comms.get_updated_coordinates(
        coordinates_for_time)]._style

    for i in range(0, total_rows_in_db):

        if sql_all_rows_data[i][1] != -1:
            ws[coordinates_for_deviceID]._style = copy(
                second_row_style["LoadWorstCase_DeviceID"])
            ws[coordinates_for_deviceID] = sql_all_rows_data[i][2]
            coordinates_for_deviceID = comms.get_updated_coordinates(
                coordinates_for_deviceID)

            ws[coordinates_for_deviceType]._style = copy(
                second_row_style["LoadWorstCase_DeviceType"])
            ws[coordinates_for_deviceType] = sql_all_rows_data[i][3]
            coordinates_for_deviceType = comms.get_updated_coordinates(
                coordinates_for_deviceType)

            if sql_all_rows_data[i][3] == "Two-winding Transformer" or sql_all_rows_data[i][3] == "Three-winding Transformer":
                ws[coordinates_for_unit]._style = copy(
                    second_row_style["LoadWorstCase_Unit"])
                ws[coordinates_for_unit] = "kVA"
                coordinates_for_unit = comms.get_updated_coordinates(
                    coordinates_for_unit)

                ws[coordinates_for_rated]._style = copy(
                    second_row_style["LoadWorstCase_Capacity"])
                ws[coordinates_for_rated] = (sql_all_rows_data[i][4])*1000
                coordinates_for_rated = comms.get_updated_coordinates(
                    coordinates_for_rated)

            else:
                ws[coordinates_for_unit]._style = copy(
                    second_row_style["LoadWorstCase_Unit"])
                ws[coordinates_for_unit] = "Amp"
                coordinates_for_unit = comms.get_updated_coordinates(
                    coordinates_for_unit)

                ws[coordinates_for_rated]._style = copy(
                    second_row_style["LoadWorstCase_Capacity"])
                ws[coordinates_for_rated] = sql_all_rows_data[i][4]
                coordinates_for_rated = comms.get_updated_coordinates(
                    coordinates_for_rated)

            ws[coordinates_for_loadingA]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhA"])
            ws[coordinates_for_loadingA] = sql_all_rows_data[i][5]
            coordinates_for_loadingA = comms.get_updated_coordinates(
                coordinates_for_loadingA)

            ws[coordinates_for_loadingB]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhB"])
            ws[coordinates_for_loadingB] = sql_all_rows_data[i][6]
            coordinates_for_loadingB = comms.get_updated_coordinates(
                coordinates_for_loadingB)

            ws[coordinates_for_loadingC]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhC"])
            ws[coordinates_for_loadingC] = sql_all_rows_data[i][7]
            coordinates_for_loadingC = comms.get_updated_coordinates(
                coordinates_for_loadingC)

            # Convert time and date to saperate columns...
            time_stamp = resultID_to_timestamp_dict[sql_all_rows_data[i][1]][0]
            date_value = time_stamp[:time_stamp.find(' ')]
            time_value = time_stamp[time_stamp.find(' ')+1:]
            date_final = datetime.strptime(date_value, '%m-%d-%Y').date()

            ws[coordinates_for_time]._style = copy(
                second_row_style["LoadWorstCase_Time"])
            ws[coordinates_for_time] = time_value
            coordinates_for_time = comms.get_updated_coordinates(
                coordinates_for_time)

            ws[coordinates_for_date]._style = copy(
                second_row_style["LoadWorstCase_Date"])
            ws[coordinates_for_date] = date_value
            coordinates_for_date = comms.get_updated_coordinates(
                coordinates_for_date)

    # Getting all data from 3 terminal devices results table
    # table_index = table_names_list.index('td3tworstoverloadcases')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TD3TWorstOverLoadCases"
    cursor.execute(command)
    sql_all_rows_data_1 = cursor.fetchall()
    total_rows_in_db_1 = len(sql_all_rows_data_1)

    db_column_name_list_1 = [description[0]
                             for description in cursor.description]

    resultID_to_timestamp_dict_1 = get_time_stamp_from_result_id(
        table_names_list, cursor)

    for i in range(0, total_rows_in_db_1):
        if sql_all_rows_data_1[i][6] != -1 or sql_all_rows_data_1[i][10] != -1 or sql_all_rows_data_1[i][14] != -1:
            # Device id for primary
            ws[coordinates_for_deviceID]._style = copy(
                second_row_style["LoadWorstCase_DeviceID"])
            ws[coordinates_for_deviceID] = sql_all_rows_data_1[i][1]
            coordinates_for_deviceID = comms.get_updated_coordinates(
                coordinates_for_deviceID)

            # Device id for secondary
            ws[coordinates_for_deviceID]._style = copy(
                second_row_style["LoadWorstCase_DeviceID"])
            ws[coordinates_for_deviceID] = sql_all_rows_data_1[i][1]
            coordinates_for_deviceID = comms.get_updated_coordinates(
                coordinates_for_deviceID)

            # Device id for tertiary
            ws[coordinates_for_deviceID]._style = copy(
                second_row_style["LoadWorstCase_DeviceID"])
            ws[coordinates_for_deviceID] = sql_all_rows_data_1[i][1]
            coordinates_for_deviceID = comms.get_updated_coordinates(
                coordinates_for_deviceID)

            # Device type for primary winding
            ws[coordinates_for_deviceType]._style = copy(
                second_row_style["LoadWorstCase_DeviceType"])
            ws[coordinates_for_deviceType] = sql_all_rows_data_1[i][2]
            coordinates_for_deviceType = comms.get_updated_coordinates(
                coordinates_for_deviceType)

            # Device type for secondary winding
            ws[coordinates_for_deviceType]._style = copy(
                second_row_style["LoadWorstCase_DeviceType"])
            ws[coordinates_for_deviceType] = sql_all_rows_data_1[i][2]
            coordinates_for_deviceType = comms.get_updated_coordinates(
                coordinates_for_deviceType)

            # Device type for tertiary winding
            ws[coordinates_for_deviceType]._style = copy(
                second_row_style["LoadWorstCase_DeviceType"])
            ws[coordinates_for_deviceType] = sql_all_rows_data_1[i][2]
            coordinates_for_deviceType = comms.get_updated_coordinates(
                coordinates_for_deviceType)

            # Unit for primary winding
            ws[coordinates_for_unit]._style = copy(
                second_row_style["LoadWorstCase_Unit"])
            ws[coordinates_for_unit] = "kVA"
            coordinates_for_unit = comms.get_updated_coordinates(
                coordinates_for_unit)

            # Unit for secondary winding
            ws[coordinates_for_unit]._style = copy(
                second_row_style["LoadWorstCase_Unit"])
            ws[coordinates_for_unit] = "kVA"
            coordinates_for_unit = comms.get_updated_coordinates(
                coordinates_for_unit)

            # Unit for tertiary winding
            ws[coordinates_for_unit]._style = copy(
                second_row_style["LoadWorstCase_Unit"])
            ws[coordinates_for_unit] = "kVA"
            coordinates_for_unit = comms.get_updated_coordinates(
                coordinates_for_unit)

            # Rated value for primary winding
            ws[coordinates_for_rated]._style = copy(
                second_row_style["LoadWorstCase_Capacity"])
            ws[coordinates_for_rated] = (sql_all_rows_data_1[i][3])*1000
            coordinates_for_rated = comms.get_updated_coordinates(
                coordinates_for_rated)

            # Rated value for secondary winding
            ws[coordinates_for_rated]._style = copy(
                second_row_style["LoadWorstCase_Capacity"])
            ws[coordinates_for_rated] = (sql_all_rows_data_1[i][4])*1000
            coordinates_for_rated = comms.get_updated_coordinates(
                coordinates_for_rated)

            # Rated value for tertiary winding
            ws[coordinates_for_rated]._style = copy(
                second_row_style["LoadWorstCase_Capacity"])
            ws[coordinates_for_rated] = (sql_all_rows_data_1[i][5])*1000
            coordinates_for_rated = comms.get_updated_coordinates(
                coordinates_for_rated)

            # Loading A for primary winding
            ws[coordinates_for_loadingA]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhA"])
            ws[coordinates_for_loadingA] = sql_all_rows_data_1[i][7]
            coordinates_for_loadingA = comms.get_updated_coordinates(
                coordinates_for_loadingA)
            # Loading B for primary winding
            ws[coordinates_for_loadingB]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhB"])
            ws[coordinates_for_loadingB] = sql_all_rows_data_1[i][8]
            coordinates_for_loadingB = comms.get_updated_coordinates(
                coordinates_for_loadingB)
            # Loading C for primary winding
            ws[coordinates_for_loadingC]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhC"])
            ws[coordinates_for_loadingC] = sql_all_rows_data_1[i][9]
            coordinates_for_loadingC = comms.get_updated_coordinates(
                coordinates_for_loadingC)

            # Loading A for secondary winding
            ws[coordinates_for_loadingA]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhA"])
            ws[coordinates_for_loadingA] = sql_all_rows_data_1[i][11]
            coordinates_for_loadingA = comms.get_updated_coordinates(
                coordinates_for_loadingA)
            # Loading B for secondary winding
            ws[coordinates_for_loadingB]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhB"])
            ws[coordinates_for_loadingB] = sql_all_rows_data_1[i][12]
            coordinates_for_loadingB = comms.get_updated_coordinates(
                coordinates_for_loadingB)
            # Loading C for secondary winding
            ws[coordinates_for_loadingC]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhC"])
            ws[coordinates_for_loadingC] = sql_all_rows_data_1[i][13]
            coordinates_for_loadingC = comms.get_updated_coordinates(
                coordinates_for_loadingC)

            # Loading A for tertiary winding
            ws[coordinates_for_loadingA]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhA"])
            ws[coordinates_for_loadingA] = sql_all_rows_data_1[i][15]
            coordinates_for_loadingA = comms.get_updated_coordinates(
                coordinates_for_loadingA)
            # Loading B for secondary winding
            ws[coordinates_for_loadingB]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhB"])
            ws[coordinates_for_loadingB] = sql_all_rows_data_1[i][16]
            coordinates_for_loadingB = comms.get_updated_coordinates(
                coordinates_for_loadingB)
            # Loading C for secondary winding
            ws[coordinates_for_loadingC]._style = copy(
                second_row_style["LoadWorstCase_MaxBranchLoadingPhC"])
            ws[coordinates_for_loadingC] = sql_all_rows_data_1[i][17]
            coordinates_for_loadingC = comms.get_updated_coordinates(
                coordinates_for_loadingC)

            # Convert time and date to saperate columns...
            # Date and time for primary winding
            time_stamp = resultID_to_timestamp_dict_1[sql_all_rows_data_1[i][6]][0]
            date_value = time_stamp[:time_stamp.find(' ')]
            time_value = time_stamp[time_stamp.find(' ')+1:]
            date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
            ws[coordinates_for_time]._style = copy(
                second_row_style["LoadWorstCase_Time"])
            ws[coordinates_for_time] = time_value
            coordinates_for_time = comms.get_updated_coordinates(
                coordinates_for_time)
            ws[coordinates_for_date]._style = copy(
                second_row_style["LoadWorstCase_Date"])
            ws[coordinates_for_date] = date_value
            coordinates_for_date = comms.get_updated_coordinates(
                coordinates_for_date)

            # Date and time for secondary winding
            time_stamp = resultID_to_timestamp_dict_1[sql_all_rows_data_1[i][10]][0]
            date_value = time_stamp[:time_stamp.find(' ')]
            time_value = time_stamp[time_stamp.find(' ')+1:]
            date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
            ws[coordinates_for_time]._style = copy(
                second_row_style["LoadWorstCase_Time"])
            ws[coordinates_for_time] = time_value
            coordinates_for_time = comms.get_updated_coordinates(
                coordinates_for_time)
            ws[coordinates_for_date]._style = copy(
                second_row_style["LoadWorstCase_Date"])
            ws[coordinates_for_date] = date_value
            coordinates_for_date = comms.get_updated_coordinates(
                coordinates_for_date)

            # Date and time for tertiary winding
            time_stamp = resultID_to_timestamp_dict_1[sql_all_rows_data_1[i][14]][0]
            date_value = time_stamp[:time_stamp.find(' ')]
            time_value = time_stamp[time_stamp.find(' ')+1:]
            date_final = datetime.strptime(date_value, '%m-%d-%Y').date()
            ws[coordinates_for_time]._style = copy(
                second_row_style["LoadWorstCase_Time"])
            ws[coordinates_for_time] = time_value
            coordinates_for_time = comms.get_updated_coordinates(
                coordinates_for_time)
            ws[coordinates_for_date]._style = copy(
                second_row_style["LoadWorstCase_Date"])
            ws[coordinates_for_date] = date_value
            coordinates_for_date = comms.get_updated_coordinates(
                coordinates_for_date)


def load_results_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    # Getting all data from system results table
    # table_index = table_names_list.index('tdsysresult')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDSysResult"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)

    db_column_name_list = [description[0]
                           for description in cursor.description]
    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    excel_column_name_to_coordinates = comms.get_excelName_to_coordinates_dict(
        wb, db_column_name_list)
    '''
    excel_column_name_to_coordinates = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, column_names):
        try:
            excel_column_name = wb.defined_names.definedName[i].name
            actual_coordinates = next(wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(actual_coordinates) if ltr == '$'][1]
            #if excel_column_name == "Results_TotalSourcekW":
            #    int_part = int(actual_coordinates[second_emp_index + 1:])+1
            #    start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            #    excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
            #elif excel_column_name == "Results_TotalSourcekvar":
            #    int_part = int(actual_coordinates[second_emp_index + 1:])+1
            #    start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            #    excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
            #elif excel_column_name == "Results_TotalLoadkW":
            #    int_part = int(actual_coordinates[second_emp_index + 1:])+1
            #    start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            #    excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
            #elif excel_column_name == "Results_TotalLoadkvar":
            #    int_part = int(actual_coordinates[second_emp_index + 1:])+1
            #    start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            #    excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
            #else:
            int_part = int(actual_coordinates[second_emp_index + 1:])+2
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue
    '''

    coordinates_for_time = excel_column_name_to_coordinates["Results_Time"]
    coordinates_for_date = excel_column_name_to_coordinates["Results_Date"]
    coordinates_for_totalkw = excel_column_name_to_coordinates["Results_TotalSourcekW"]
    coordinates_for_totalkvar = excel_column_name_to_coordinates["Results_TotalSourcekvar"]
    coordinates_for_windkw = excel_column_name_to_coordinates["Results_TotalWindkW"]
    coordinates_for_solarkw = excel_column_name_to_coordinates["Results_TotalSolarkW"]
    coordinates_for_dcSourcekw = excel_column_name_to_coordinates["Results_TotalDCSourcekW"]
    coordinates_for_totalLoadkw = excel_column_name_to_coordinates["Results_TotalLoadkW"]
    coordinates_for_totalLoadkvar = excel_column_name_to_coordinates["Results_TotalLoadkvar"]
    coordinates_for_totalDCLoadkw = excel_column_name_to_coordinates["Results_TotalDCLoadkW"]
    coordinates_for_totalLosseskW = excel_column_name_to_coordinates["Results_TotalLosseskW"]

    # Storing style of each cell in the second row in a dictionary.
    second_row_style = dict()  # {"range_string":cell_style}
    second_row_style["Results_Time"] = ws[comms.get_updated_coordinates(
        coordinates_for_time)]._style
    second_row_style["Results_Date"] = ws[comms.get_updated_coordinates(
        coordinates_for_date)]._style
    second_row_style["Results_TotalSourcekW"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalkw)]._style
    second_row_style["Results_TotalSourcekvar"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalkvar)]._style
    second_row_style["Results_TotalWindkW"] = ws[comms.get_updated_coordinates(
        coordinates_for_windkw)]._style
    second_row_style["Results_TotalSolarkW"] = ws[comms.get_updated_coordinates(
        coordinates_for_solarkw)]._style
    second_row_style["Results_TotalDCSourcekW"] = ws[comms.get_updated_coordinates(
        coordinates_for_dcSourcekw)]._style
    second_row_style["Results_TotalLoadkW"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalLoadkw)]._style
    second_row_style["Results_TotalLoadkvar"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalLoadkvar)]._style
    second_row_style["Results_TotalDCLoadkW"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalDCLoadkw)]._style
    second_row_style["Results_TotalLosseskW"] = ws[comms.get_updated_coordinates(
        coordinates_for_totalLosseskW)]._style

    for i in range(0, total_rows_in_db):
        # Convert time and date to saperate columns...
        time_stamp = resultID_to_timestamp_dict[sql_all_rows_data[i][1]][0]
        date_value = time_stamp[:time_stamp.find(' ')]
        time_value = time_stamp[time_stamp.find(' ')+1:]
        date_final = datetime.strptime(date_value, '%m-%d-%Y').date()

        ws[coordinates_for_time]._style = copy(
            second_row_style["Results_Time"])
        ws[coordinates_for_time] = time_value
        coordinates_for_time = comms.get_updated_coordinates(
            coordinates_for_time)

        ws[coordinates_for_date]._style = copy(
            second_row_style["Results_Date"])
        ws[coordinates_for_date] = date_value
        coordinates_for_date = comms.get_updated_coordinates(
            coordinates_for_date)

        total_ac_load_kw = (
            (sql_all_rows_data[i][2] + sql_all_rows_data[i][3] + sql_all_rows_data[i][4])*1000)
        ws[coordinates_for_totalLoadkw]._style = copy(
            second_row_style["Results_TotalLoadkW"])
        ws[coordinates_for_totalLoadkw] = total_ac_load_kw
        coordinates_for_totalLoadkw = comms.get_updated_coordinates(
            coordinates_for_totalLoadkw)

        total_ac_load_kvar = (
            (sql_all_rows_data[i][5] + sql_all_rows_data[i][6] + sql_all_rows_data[i][7])*1000)
        ws[coordinates_for_totalLoadkvar]._style = copy(
            second_row_style["Results_TotalLoadkvar"])
        ws[coordinates_for_totalLoadkvar] = total_ac_load_kvar
        coordinates_for_totalLoadkvar = comms.get_updated_coordinates(
            coordinates_for_totalLoadkvar)

        total_ac_source_kw = (
            (sql_all_rows_data[i][8] + sql_all_rows_data[i][9] + sql_all_rows_data[i][10])*1000)
        ws[coordinates_for_totalkw]._style = copy(
            second_row_style["Results_TotalSourcekW"])
        ws[coordinates_for_totalkw] = total_ac_source_kw
        coordinates_for_totalkw = comms.get_updated_coordinates(
            coordinates_for_totalkw)

        total_ac_source_kvar = (
            (sql_all_rows_data[i][11] + sql_all_rows_data[i][12] + sql_all_rows_data[i][13])*1000)
        ws[coordinates_for_totalkvar]._style = copy(
            second_row_style["Results_TotalSourcekvar"])
        ws[coordinates_for_totalkvar] = total_ac_source_kvar
        coordinates_for_totalkvar = comms.get_updated_coordinates(
            coordinates_for_totalkvar)

        total_dc_load = (sql_all_rows_data[i][70]*1000)
        ws[coordinates_for_totalDCLoadkw]._style = copy(
            second_row_style["Results_TotalDCLoadkW"])
        ws[coordinates_for_totalDCLoadkw] = total_dc_load
        coordinates_for_totalDCLoadkw = comms.get_updated_coordinates(
            coordinates_for_totalDCLoadkw)

        total_dc_source = (sql_all_rows_data[i][71]*1000)
        ws[coordinates_for_dcSourcekw]._style = copy(
            second_row_style["Results_TotalDCSourcekW"])
        ws[coordinates_for_dcSourcekw] = total_dc_source
        coordinates_for_dcSourcekw = comms.get_updated_coordinates(
            coordinates_for_dcSourcekw)

        total_ac_dc_source_kw = total_ac_source_kw + total_dc_source
        total_ac_dc_load_kw = total_ac_load_kw + total_dc_load
        total_losses = total_ac_dc_source_kw - total_ac_dc_load_kw
        ws[coordinates_for_totalLosseskW]._style = copy(
            second_row_style["Results_TotalLosseskW"])
        ws[coordinates_for_totalLosseskW] = total_losses
        coordinates_for_totalLosseskW = comms.get_updated_coordinates(
            coordinates_for_totalLosseskW)

        ws[coordinates_for_windkw]._style = copy(
            second_row_style["Results_TotalWindkW"])
        ws[coordinates_for_windkw] = (sql_all_rows_data[i][72]*1000)
        coordinates_for_windkw = comms.get_updated_coordinates(
            coordinates_for_windkw)

        ws[coordinates_for_solarkw]._style = copy(
            second_row_style["Results_TotalSolarkW"])
        ws[coordinates_for_solarkw] = (sql_all_rows_data[i][73]*1000)
        coordinates_for_solarkw = comms.get_updated_coordinates(
            coordinates_for_solarkw)


def load_voltage_worst_case_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    # Getting all data from system results table
    # table_index = table_names_list.index('tdworstvoltagecases')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDWorstVoltageCases"
    cursor.execute(command)
    sql_all_rows_data = cursor.fetchall()
    total_rows_in_db = len(sql_all_rows_data)
    db_column_name_list = [description[0]
                           for description in cursor.description]
    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    excel_column_name_to_coordinates = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, column_names):
        try:
            excel_column_name = wb.defined_names.definedName[i].name
            actual_coordinates = next(
                wb.defined_names[excel_column_name].destinations)[1]
            second_emp_index = [i for i, ltr in enumerate(
                actual_coordinates) if ltr == '$'][1]
            int_part = int(actual_coordinates[second_emp_index + 1:])+1
            start_writing_coordinate = actual_coordinates[:second_emp_index+1] + str(
                int_part)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    coordinates_for_deviceID = excel_column_name_to_coordinates["PhaseVoltageWorstCase_DeviceID"]
    coordinates_for_nomkV = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_DeviceNominalKV"]
    coordinates_for_connectionType = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_ConnectionType"]
    coordinates_for_maxVoltPhase = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MaxVoltPhase"]
    coordinates_for_maxVoltValue = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MaxVoltValue"]
    coordinates_for_maxVoltDate = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MaxVoltDate"]
    coordinates_for_maxVoltTime = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MaxVoltTime"]
    coordinates_for_minVoltPhase = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MinVoltPhase"]
    coordinates_for_minVoltValue = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MinVoltValue"]
    coordinates_for_minVoltDate = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MinVoltDate"]
    coordinates_for_minVoltTime = excel_column_name_to_coordinates[
        "PhaseVoltageWorstCase_MinVoltTime"]

    # Storing style of each cell in the second row in a dictionary.
    second_row_style = dict()  # {"range_string":cell_style}
    second_row_style["PhaseVoltageWorstCase_DeviceID"] = ws[comms.get_updated_coordinates(
        coordinates_for_deviceID)]._style
    second_row_style["PhaseVoltageWorstCase_DeviceNominalKV"] = ws[comms.get_updated_coordinates(
        coordinates_for_nomkV)]._style
    second_row_style["PhaseVoltageWorstCase_ConnectionType"] = ws[comms.get_updated_coordinates(
        coordinates_for_connectionType)]._style
    second_row_style["PhaseVoltageWorstCase_MaxVoltPhase"] = ws[comms.get_updated_coordinates(
        coordinates_for_maxVoltPhase)]._style
    second_row_style["PhaseVoltageWorstCase_MaxVoltValue"] = ws[comms.get_updated_coordinates(
        coordinates_for_maxVoltValue)]._style
    second_row_style["PhaseVoltageWorstCase_MaxVoltDate"] = ws[comms.get_updated_coordinates(
        coordinates_for_maxVoltDate)]._style
    second_row_style["PhaseVoltageWorstCase_MaxVoltTime"] = ws[comms.get_updated_coordinates(
        coordinates_for_maxVoltTime)]._style
    second_row_style["PhaseVoltageWorstCase_MinVoltPhase"] = ws[comms.get_updated_coordinates(
        coordinates_for_minVoltPhase)]._style
    second_row_style["PhaseVoltageWorstCase_MinVoltValue"] = ws[comms.get_updated_coordinates(
        coordinates_for_minVoltValue)]._style
    second_row_style["PhaseVoltageWorstCase_MinVoltDate"] = ws[comms.get_updated_coordinates(
        coordinates_for_minVoltDate)]._style
    second_row_style["PhaseVoltageWorstCase_MinVoltTime"] = ws[comms.get_updated_coordinates(
        coordinates_for_minVoltTime)]._style

    for i in range(0, total_rows_in_db):

        # Only adding data without result ID less than 0
        if sql_all_rows_data[i][15] > 0 or sql_all_rows_data[i][11] > 0:
            ws[coordinates_for_deviceID]._style = copy(
                second_row_style["PhaseVoltageWorstCase_DeviceID"])
            ws[coordinates_for_deviceID] = sql_all_rows_data[i][1]
            coordinates_for_deviceID = comms.get_updated_coordinates(
                coordinates_for_deviceID)

            ws[coordinates_for_nomkV]._style = copy(
                second_row_style["PhaseVoltageWorstCase_DeviceNominalKV"])
            ws[coordinates_for_nomkV] = sql_all_rows_data[i][2]
            coordinates_for_nomkV = comms.get_updated_coordinates(
                coordinates_for_nomkV)

            # Followig data is not available in database, thus defaulting value to LN
            ws[coordinates_for_connectionType]._style = copy(
                second_row_style["PhaseVoltageWorstCase_ConnectionType"])
            ws[coordinates_for_connectionType] = "LN"
            coordinates_for_connectionType = comms.get_updated_coordinates(
                coordinates_for_connectionType)

            # writing Max phase voltage values
            if sql_all_rows_data[i][11] > 0:

                max_voltage_list = []

                max_voltage_list.append(sql_all_rows_data[i][12])
                max_voltage_list.append(sql_all_rows_data[i][13])
                max_voltage_list.append(sql_all_rows_data[i][14])

                max_voltage = max(max_voltage_list)
                max_voltage_index = max_voltage_list.index(max_voltage)

                ws[coordinates_for_maxVoltPhase]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MaxVoltPhase"])
                if max_voltage_index == 0:
                    ws[coordinates_for_maxVoltPhase] = "A"

                if max_voltage_index == 1:
                    ws[coordinates_for_maxVoltPhase] = "B"

                if max_voltage_index == 2:
                    ws[coordinates_for_maxVoltPhase] = "C"
                coordinates_for_maxVoltPhase = comms.get_updated_coordinates(
                    coordinates_for_maxVoltPhase)

                # adding voltage value.
                ws[coordinates_for_maxVoltValue]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MaxVoltValue"])
                ws[coordinates_for_maxVoltValue] = max_voltage/100
                coordinates_for_maxVoltValue = comms.get_updated_coordinates(
                    coordinates_for_maxVoltValue)

                # Convert time and date to saperate columns. Adding max value date and time
                time_stamp = resultID_to_timestamp_dict[sql_all_rows_data[i][11]][0]
                date_value = time_stamp[:time_stamp.find(' ')]
                time_value = time_stamp[time_stamp.find(' ')+1:]
                date_final = datetime.strptime(date_value, '%m-%d-%Y').date()

                ws[coordinates_for_maxVoltTime]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MaxVoltTime"])
                ws[coordinates_for_maxVoltTime] = time_value
                coordinates_for_maxVoltTime = comms.get_updated_coordinates(
                    coordinates_for_maxVoltTime)

                ws[coordinates_for_maxVoltDate]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MaxVoltDate"])
                ws[coordinates_for_maxVoltDate] = date_value
                coordinates_for_maxVoltDate = comms.get_updated_coordinates(
                    coordinates_for_maxVoltDate)

            # writing Min phase voltage values
            if sql_all_rows_data[i][15] > 0:

                min_voltage_list = []

                min_voltage_list.append(sql_all_rows_data[i][16])
                min_voltage_list.append(sql_all_rows_data[i][17])
                min_voltage_list.append(sql_all_rows_data[i][18])

                min_voltage = min(min_voltage_list)
                min_voltage_index = min_voltage_list.index(min_voltage)

                ws[coordinates_for_minVoltPhase]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MinVoltPhase"])
                if min_voltage_index == 0:
                    ws[coordinates_for_minVoltPhase] = "A"

                if min_voltage_index == 1:
                    ws[coordinates_for_minVoltPhase] = "B"

                if min_voltage_index == 2:
                    ws[coordinates_for_minVoltPhase] = "C"
                coordinates_for_minVoltPhase = comms.get_updated_coordinates(
                    coordinates_for_minVoltPhase)

                # adding voltage value.
                ws[coordinates_for_minVoltValue]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MinVoltValue"])
                ws[coordinates_for_minVoltValue] = min_voltage/100
                coordinates_for_minVoltValue = comms.get_updated_coordinates(
                    coordinates_for_minVoltValue)

                # Convert time and date to saperate columns. Adding max value date and time
                time_stamp = resultID_to_timestamp_dict[sql_all_rows_data[i][15]][0]
                date_value = time_stamp[:time_stamp.find(' ')]
                time_value = time_stamp[time_stamp.find(' ')+1:]
                date_final = datetime.strptime(date_value, '%m-%d-%Y').date()

                ws[coordinates_for_minVoltTime]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MinVoltTime"])
                ws[coordinates_for_minVoltTime] = time_value
                coordinates_for_minVoltTime = comms.get_updated_coordinates(
                    coordinates_for_minVoltTime)

                ws[coordinates_for_minVoltDate]._style = copy(
                    second_row_style["PhaseVoltageWorstCase_MinVoltDate"])
                ws[coordinates_for_minVoltDate] = date_value
                coordinates_for_minVoltDate = comms.get_updated_coordinates(
                    coordinates_for_minVoltDate)


def load_summary_group_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Getting all data from group info table
    # table_index = table_names_list.index('tdgroupinfo')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDGroupInfo"
    cursor.execute(command)
    group_info_data = cursor.fetchall()
    total_rows_group_info = len(group_info_data)
    group_info_column_name_list = [description[0]
                                   for description in cursor.description]

    # Getting all data from group results table
    # table_index = table_names_list.index('tdgroupresult')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDGroupResult"
    cursor.execute(command)
    group_result_data = cursor.fetchall()
    total_rows_group_result = len(group_result_data)
    group_result_column_name_list = [description[0]
                                     for description in cursor.description]
    column_names = len(wb.defined_names.definedName)

    # Getting train study case info table
    # table_index = table_names_list.index('tdstudycaseinfo')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDStudyCaseInfo"
    cursor.execute(command)
    study_case_info_data = cursor.fetchall()
    total_rows_study_case_info = len(study_case_info_data)
    study_case_info_column_name_list = [
        description[0] for description in cursor.description]

    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    excel_column_name_to_coordinates = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, column_names):
        try:
            excel_column_name = wb.defined_names.definedName[i].name
            actual_coordinates = next(
                wb.defined_names[excel_column_name].destinations)[1]
            first_emp_index = [i for i, ltr in enumerate(
                actual_coordinates) if ltr == '$'][0]
            second_emp_index = [i for i, ltr in enumerate(
                actual_coordinates) if ltr == '$'][1]
            column_alphabet = actual_coordinates[first_emp_index +
                                                 1:second_emp_index]
            if len(column_alphabet) == 1:
                if column_alphabet == 'Z':
                    start_writing_column = 'AA'
                else:
                    start_writing_column = chr(ord(column_alphabet)+1)

            elif len(column_alphabet) == 2:
                if column_alphabet == 'ZZ':
                    start_writing_column = 'AAA'
                else:
                    first_char_of_column_alphabet = column_alphabet[:len(
                        column_alphabet)-1]
                    second_char_of_column_alphabet = column_alphabet[len(
                        column_alphabet)-1:]
                    start_writing_column = first_char_of_column_alphabet + \
                        chr(ord(second_char_of_column_alphabet)+1)

            start_writing_coordinate = actual_coordinates[:first_emp_index+1] + \
                start_writing_column + actual_coordinates[second_emp_index:]
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    # Collecting number of Groups in tdgroupinfo table
    number_of_groups = total_rows_group_info
    end_time_stamp = datetime.strptime(
        (study_case_info_data[0][2]), '%m-%d-%Y %H:%M:%S')

    for group in range(0, number_of_groups):
        ws = wb['GroupSummary']
        # Writing data from group info table....
        coordinates_for_groupType = excel_column_name_to_coordinates["GroupSummary_GroupType"]
        ws[coordinates_for_groupType] = group_info_data[group][2]
        coordinates_for_groupName = excel_column_name_to_coordinates["GroupSummary_GroupName"]
        ws[coordinates_for_groupName] = group_info_data[group][3]
        group_id = group_info_data[group][1]

        data_for_current_group = []
        for i in range(0, total_rows_group_result):
            if group_result_data[i][2] == group_id:
                data_for_current_group.append(group_result_data[i])

        list_of_AC_total_gen = []
        AC_total_gen_to_time_stamp = dict()
        total_AC_generation_energy = 0

        list_of_AC_total_load = []
        AC_total_load_to_time_stamp = dict()
        total_AC_load_energy = 0

        list_of_DC_total_gen = []
        DC_total_gen_to_time_stamp = dict()
        total_DC_generation_energy = 0

        list_of_DC_total_load = []
        DC_total_load_to_time_stamp = dict()
        total_DC_load_energy = 0

        list_of_wind_gen = []
        wind_gen_to_time_stamp = dict()
        total_wind_energy = 0

        list_of_solar_gen = []
        solar_gen_to_time_stamp = dict()
        total_solar_energy = 0

        # Generating data from group result table:
        for i in range(0, len(data_for_current_group)):

            time_stamp_for_i = datetime.strptime(
                resultID_to_timestamp_dict[data_for_current_group[i][1]][0], '%m-%d-%Y %H:%M:%S')

            if(i == len(data_for_current_group)-1):
                if(end_time_stamp != datetime.strptime(resultID_to_timestamp_dict[data_for_current_group[i][1]][0], '%m-%d-%Y %H:%M:%S')):
                    time_stamp_for_i1 = end_time_stamp
                else:
                    time_stamp_for_i1 = datetime.strptime(
                        resultID_to_timestamp_dict[data_for_current_group[i][1]][0], '%m-%d-%Y %H:%M:%S')
            else:
                time_stamp_for_i1 = datetime.strptime(
                    resultID_to_timestamp_dict[data_for_current_group[i+1][1]][0], '%m-%d-%Y %H:%M:%S')

            duration = (time_stamp_for_i1 - time_stamp_for_i)
            duration_in_hours = (time_stamp_for_i1 -
                                 time_stamp_for_i).total_seconds()/3600

            total_AC_generation_for_i = data_for_current_group[i][9] + \
                data_for_current_group[i][10] + data_for_current_group[i][11]
            total_AC_generation_energy += total_AC_generation_for_i * duration_in_hours
            list_of_AC_total_gen.append(total_AC_generation_for_i)
            AC_total_gen_to_time_stamp[total_AC_generation_for_i] = time_stamp_for_i

            total_solar_generation_for_i = data_for_current_group[i][24]
            total_solar_energy += total_solar_generation_for_i * duration_in_hours
            list_of_solar_gen.append(total_solar_generation_for_i)
            solar_gen_to_time_stamp[total_solar_generation_for_i] = time_stamp_for_i

            total_wind_generation_for_i = data_for_current_group[i][23]
            total_wind_energy += total_wind_generation_for_i * duration_in_hours
            list_of_wind_gen.append(total_wind_generation_for_i)
            wind_gen_to_time_stamp[total_wind_generation_for_i] = time_stamp_for_i

            total_DC_generation_for_i = data_for_current_group[i][22]
            total_DC_generation_energy += total_DC_generation_for_i * duration_in_hours
            list_of_DC_total_gen.append(total_DC_generation_for_i)
            DC_total_gen_to_time_stamp[total_DC_generation_for_i] = time_stamp_for_i

            total_AC_load_for_i = data_for_current_group[i][3] + \
                data_for_current_group[i][4] + data_for_current_group[i][5]
            total_AC_load_energy += total_AC_load_for_i * duration_in_hours
            list_of_AC_total_load.append(total_AC_load_for_i)
            AC_total_load_to_time_stamp[total_AC_load_for_i] = time_stamp_for_i

            total_DC_load_for_i = data_for_current_group[i][21]
            total_DC_load_energy += total_DC_load_for_i * duration_in_hours
            list_of_DC_total_load.append(total_DC_load_for_i)
            DC_total_load_to_time_stamp[total_DC_load_for_i] = time_stamp_for_i

        total_system_generation_energy = total_AC_generation_energy + \
            total_solar_energy + total_wind_energy + total_DC_generation_energy
        total_system_load_energy = total_AC_load_energy + total_DC_load_energy

        #total_energy_loss = total_system_generation_energy - total_system_load_energy

        # Calculating Load Factor
        average_demand = (sum(list_of_DC_total_load) +
                          sum(list_of_AC_total_load)) / len(data_for_current_group)
        ac_plus_dc_load_list = list_of_DC_total_load + list_of_AC_total_load
        max_demand = max(ac_plus_dc_load_list)

        if max_demand != 0:
            load_factor = average_demand/max_demand
        else:
            load_factor = -1
        #load_factor = average_demand/max_demand

        # Writing data for MAX Power section
        coordinates_for_maxWindGen = excel_column_name_to_coordinates["GroupSummary_MaxWindGen"]
        coordinates_for_maxWindGen_date = comms.increase_column_by_one(
            coordinates_for_maxWindGen)
        coordinates_for_maxWindGen_time = comms.increase_column_by_one(
            coordinates_for_maxWindGen_date)
        time_stamp_for_maxWindGen = wind_gen_to_time_stamp[max(
            list_of_wind_gen)]
        ws[coordinates_for_maxWindGen] = max(list_of_wind_gen)
        ws[coordinates_for_maxWindGen_date] = time_stamp_for_maxWindGen.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxWindGen_time] = time_stamp_for_maxWindGen.time().strftime('%X')

        coordinates_for_maxSolarGen = excel_column_name_to_coordinates["GroupSummary_MaxSolarGen"]
        coordinates_for_maxSolarGen_date = comms.increase_column_by_one(
            coordinates_for_maxSolarGen)
        coordinates_for_maxSolarGen_time = comms.increase_column_by_one(
            coordinates_for_maxSolarGen_date)
        time_stamp_for_maxSolarGen = solar_gen_to_time_stamp[max(
            list_of_solar_gen)]
        ws[coordinates_for_maxSolarGen] = max(list_of_solar_gen)
        ws[coordinates_for_maxSolarGen_date] = time_stamp_for_maxSolarGen.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxSolarGen_time] = time_stamp_for_maxSolarGen.time(
        ).strftime('%X')

        coordinates_for_maxTotalACGen = excel_column_name_to_coordinates["GroupSummary_MaxACGen"]
        coordinates_for_maxTotalACGen_date = comms.increase_column_by_one(
            coordinates_for_maxTotalACGen)
        coordinates_for_maxTotalACGen_time = comms.increase_column_by_one(
            coordinates_for_maxTotalACGen_date)
        time_stamp_for_maxTotalACGen = AC_total_gen_to_time_stamp[max(
            list_of_AC_total_gen)]
        ws[coordinates_for_maxTotalACGen] = max(list_of_AC_total_gen)
        ws[coordinates_for_maxTotalACGen_date] = time_stamp_for_maxTotalACGen.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxTotalACGen_time] = time_stamp_for_maxTotalACGen.time(
        ).strftime('%X')

        coordinates_for_maxTotalDCGen = excel_column_name_to_coordinates["GroupSummary_MaxDCGen"]
        coordinates_for_maxTotalDCGen_date = comms.increase_column_by_one(
            coordinates_for_maxTotalDCGen)
        coordinates_for_maxTotalDCGen_time = comms.increase_column_by_one(
            coordinates_for_maxTotalDCGen_date)
        time_stamp_for_maxTotalDCGen = DC_total_gen_to_time_stamp[max(
            list_of_DC_total_gen)]
        ws[coordinates_for_maxTotalDCGen] = max(list_of_DC_total_gen)
        ws[coordinates_for_maxTotalDCGen_date] = time_stamp_for_maxTotalDCGen.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxTotalDCGen_time] = time_stamp_for_maxTotalDCGen.time(
        ).strftime('%X')

        coordinates_for_maxTotalACLoad = excel_column_name_to_coordinates[
            "GroupSummary_MaxACLoad"]
        coordinates_for_maxTotalACLoad_date = comms.increase_column_by_one(
            coordinates_for_maxTotalACLoad)
        coordinates_for_maxTotalACLoad_time = comms.increase_column_by_one(
            coordinates_for_maxTotalACLoad_date)
        time_stamp_for_maxTotalACLoad = AC_total_load_to_time_stamp[max(
            list_of_AC_total_load)]
        ws[coordinates_for_maxTotalACLoad] = max(list_of_AC_total_load)
        ws[coordinates_for_maxTotalACLoad_date] = time_stamp_for_maxTotalACLoad.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxTotalACLoad_time] = time_stamp_for_maxTotalACLoad.time(
        ).strftime('%X')

        coordinates_for_maxTotalDCLoad = excel_column_name_to_coordinates[
            "GroupSummary_MaxDCLoad"]
        coordinates_for_maxTotalDCLoad_date = comms.increase_column_by_one(
            coordinates_for_maxTotalDCLoad)
        coordinates_for_maxTotalDCLoad_time = comms.increase_column_by_one(
            coordinates_for_maxTotalDCLoad_date)
        time_stamp_for_maxTotalDCLoad = DC_total_load_to_time_stamp[max(
            list_of_DC_total_load)]
        ws[coordinates_for_maxTotalDCLoad] = max(list_of_DC_total_load)
        ws[coordinates_for_maxTotalDCLoad_date] = time_stamp_for_maxTotalDCLoad.strftime(
            '%m-%d-%Y')
        ws[coordinates_for_maxTotalDCLoad_time] = time_stamp_for_maxTotalDCLoad.time(
        ).strftime('%X')

        # Writing calculate data for "Total Enegry" fields in excel
        coordinates_for_windGen = excel_column_name_to_coordinates["GroupSummary_TotalWindGen"]
        ws[coordinates_for_windGen] = total_wind_energy

        coordinates_for_solarGen = excel_column_name_to_coordinates["GroupSummary_TotalSolarGen"]
        ws[coordinates_for_solarGen] = total_solar_energy

        coordinates_for_ACGen = excel_column_name_to_coordinates["GroupSummary_TotalACGen"]
        ws[coordinates_for_ACGen] = total_AC_generation_energy

        coordinates_for_DCGen = excel_column_name_to_coordinates["GroupSummary_TotalDCGen"]
        ws[coordinates_for_DCGen] = total_DC_generation_energy

        coordinates_for_sysGen = excel_column_name_to_coordinates["GroupSummary_TotalSysGen"]
        ws[coordinates_for_sysGen] = total_system_generation_energy

        coordinates_for_ACLoad = excel_column_name_to_coordinates["GroupSummary_TotalACLoad"]
        ws[coordinates_for_ACLoad] = total_AC_load_energy

        coordinates_for_DCLoad = excel_column_name_to_coordinates["GroupSummary_TotalDCLoad"]
        ws[coordinates_for_DCLoad] = total_DC_load_energy

        coordinates_for_sysLoad = excel_column_name_to_coordinates["GroupSummary_TotalSysLoad"]
        ws[coordinates_for_sysLoad] = total_system_load_energy

        #coordinates_for_loss = excel_column_name_to_coordinates["GroupSummary_TotalLoss"]
        #ws[coordinates_for_loss] = total_energy_loss

        coordinates_for_loadfactor = excel_column_name_to_coordinates["GroupSummary_LoadFactor"]

        if load_factor != -1:
            ws[coordinates_for_loadfactor] = load_factor
        else:
            ws[coordinates_for_loadfactor] = 'N/A'
        #ws[coordinates_for_loadfactor] = load_factor

        ws_new = wb.copy_worksheet(ws)
        groupName = group_info_data[group][3]
        ws_new.title = 'Group Summary-' + groupName
        ws_new.column_dimensions['A'].width = 1.65

    wb.remove_sheet(wb["GroupSummary"])


def load_system_summary_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    def get_time_stamp_from_result_id(table_names_list, cursor):
        # table_index = table_names_list.index('tdtimeid')
        # command = "SELECT * FROM " + table_names_list[table_index]
        command = "SELECT * FROM TDTimeID"
        cursor.execute(command)
        sql_all_rows_data = cursor.fetchall()

        result_id_to_timestamp_dict = dict()
        result_id_list = []
        timestamp_list = []

        for i in range(len(sql_all_rows_data)):
            result_id_list.append(sql_all_rows_data[i][3])
            timestamp_list.append(sql_all_rows_data[i][2])

        timestamp_resultID_list = zip(result_id_list, timestamp_list)
        resultID_to_timestamp_dict = collections.defaultdict(list)
        for v, k in timestamp_resultID_list:
            resultID_to_timestamp_dict[v].append(k)
        return resultID_to_timestamp_dict

    resultID_to_timestamp_dict = get_time_stamp_from_result_id(
        table_names_list, cursor)

    # Getting all data from system results table
    # table_index = table_names_list.index('tdsysresult')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDSysResult"
    cursor.execute(command)
    system_result_data = cursor.fetchall()
    total_rows_system_result = len(system_result_data)
    group_result_column_name_list = [description[0]
                                     for description in cursor.description]

    # Getting train study case info table
    # table_index = table_names_list.index('tdstudycaseinfo')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDStudyCaseInfo"
    cursor.execute(command)
    study_case_info_data = cursor.fetchall()
    total_rows_study_case_info = len(study_case_info_data)
    study_case_info_column_name_list = [
        description[0] for description in cursor.description]

    # Connecting to Excel and Fetching Defined Names
    column_names = len(wb.defined_names.definedName)
    index_of_underscore = wb.defined_names.definedName[1].name.find("_")
    excel_column_name_prefix = wb.defined_names.definedName[1].name[:index_of_underscore+1]

    excel_column_name_to_coordinates = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, column_names):
        try:
            excel_column_name = wb.defined_names.definedName[i].name
            actual_coordinates = next(
                wb.defined_names[excel_column_name].destinations)[1]
            start_writing_coordinate = comms.increase_column_by_one(
                actual_coordinates)
            excel_column_name_to_coordinates[excel_column_name] = start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    list_of_AC_total_gen = []
    AC_total_gen_to_time_stamp = dict()
    total_AC_generation_energy = 0

    list_of_AC_total_load = []
    AC_total_load_to_time_stamp = dict()
    total_AC_load_energy = 0

    list_of_DC_total_gen = []
    DC_total_gen_to_time_stamp = dict()
    total_DC_generation_energy = 0

    list_of_DC_total_load = []
    DC_total_load_to_time_stamp = dict()
    total_DC_load_energy = 0

    list_of_wind_gen = []
    wind_gen_to_time_stamp = dict()
    total_wind_energy = 0

    list_of_solar_gen = []
    solar_gen_to_time_stamp = dict()
    total_solar_energy = 0

    Va_max_list = []
    Va_max_to_resultID_busID = dict()
    Vb_max_list = []
    Vb_max_to_resultID_busID = dict()
    Vc_max_list = []
    Vc_max_to_resultID_busID = dict()

    Va_min_list = []
    Va_min_to_resultID_busID = dict()
    Vb_min_list = []
    Vb_min_to_resultID_busID = dict()
    Vc_min_list = []
    Vc_min_to_resultID_busID = dict()
    end_time_stamp = datetime.strptime(
        (study_case_info_data[0][2]), '%m-%d-%Y %H:%M:%S')
    # Generating data from system result table:
    for j in range(0, total_rows_system_result):
        Va_max_list.append(system_result_data[j][14])
        Va_max_to_resultID_busID[system_result_data[j][14]] = [
            system_result_data[j][1], system_result_data[j][18]]
        Va_min_list.append(system_result_data[j][24])
        Va_min_to_resultID_busID[system_result_data[j][24]] = [
            system_result_data[j][1], system_result_data[j][28]]

        Vb_max_list.append(system_result_data[j][15])
        Vb_max_to_resultID_busID[system_result_data[j][15]] = [
            system_result_data[j][1], system_result_data[j][18]]
        Vb_min_list.append(system_result_data[j][25])
        Vb_min_to_resultID_busID[system_result_data[j][25]] = [
            system_result_data[j][1], system_result_data[j][28]]

        Vc_max_list.append(system_result_data[j][16])
        Vc_max_to_resultID_busID[system_result_data[j][16]] = [
            system_result_data[j][1], system_result_data[j][18]]
        Vc_min_list.append(system_result_data[j][26])
        Vc_min_to_resultID_busID[system_result_data[j][26]] = [
            system_result_data[j][1], system_result_data[j][28]]

    # Generating data for energy calculation table:
    for i in range(0, total_rows_system_result):
        time_stamp_for_i = datetime.strptime(
            resultID_to_timestamp_dict[system_result_data[i][1]][0], '%m-%d-%Y %H:%M:%S')

        if(i == total_rows_system_result-1):
            if(end_time_stamp != datetime.strptime(resultID_to_timestamp_dict[system_result_data[i][1]][0], '%m-%d-%Y %H:%M:%S')):
                time_stamp_for_i1 = end_time_stamp
            else:
                time_stamp_for_i1 = datetime.strptime(
                    resultID_to_timestamp_dict[system_result_data[i][1]][0], '%m-%d-%Y %H:%M:%S')
        else:
            time_stamp_for_i1 = datetime.strptime(
                resultID_to_timestamp_dict[system_result_data[i+1][1]][0], '%m-%d-%Y %H:%M:%S')

        duration = (time_stamp_for_i1 - time_stamp_for_i)
        duration_in_hours = (time_stamp_for_i1 -
                             time_stamp_for_i).total_seconds()/3600

        total_AC_generation_for_i = system_result_data[i][8] + \
            system_result_data[i][9] + system_result_data[i][10]
        total_AC_generation_energy += total_AC_generation_for_i * duration_in_hours
        list_of_AC_total_gen.append(total_AC_generation_for_i)
        AC_total_gen_to_time_stamp[total_AC_generation_for_i] = time_stamp_for_i

        total_solar_generation_for_i = system_result_data[i][73]
        total_solar_energy += total_solar_generation_for_i * duration_in_hours
        list_of_solar_gen.append(total_solar_generation_for_i)
        solar_gen_to_time_stamp[total_solar_generation_for_i] = time_stamp_for_i

        total_wind_generation_for_i = system_result_data[i][72]
        total_wind_energy += total_wind_generation_for_i * duration_in_hours
        list_of_wind_gen.append(total_wind_generation_for_i)
        wind_gen_to_time_stamp[total_wind_generation_for_i] = time_stamp_for_i

        total_DC_generation_for_i = system_result_data[i][71]
        total_DC_generation_energy += total_DC_generation_for_i * duration_in_hours
        list_of_DC_total_gen.append(total_DC_generation_for_i)
        DC_total_gen_to_time_stamp[total_DC_generation_for_i] = time_stamp_for_i

        total_AC_load_for_i = system_result_data[i][2] + \
            system_result_data[i][3] + system_result_data[i][4]
        total_AC_load_energy += total_AC_load_for_i * duration_in_hours
        list_of_AC_total_load.append(total_AC_load_for_i)
        AC_total_load_to_time_stamp[total_AC_load_for_i] = time_stamp_for_i

        total_DC_load_for_i = system_result_data[i][70]
        total_DC_load_energy += total_DC_load_for_i * duration_in_hours
        list_of_DC_total_load.append(total_DC_load_for_i)
        DC_total_load_to_time_stamp[total_DC_load_for_i] = time_stamp_for_i

    total_system_generation_energy = total_AC_generation_energy + \
        total_solar_energy + total_wind_energy + total_DC_generation_energy
    total_system_load_energy = total_AC_load_energy + total_DC_load_energy

    total_energy_loss = total_system_generation_energy - total_system_load_energy

    # Calculation for Load Factor
    total_demand = (sum(list_of_DC_total_load) + sum(list_of_AC_total_load)) / \
        total_rows_system_result  # replace len with total time steps
    ac_demand_plus_dc_demand_list = list_of_DC_total_load + list_of_AC_total_load
    # max of AC and DC load together = max_demand
    max_demand = max(ac_demand_plus_dc_demand_list)
    # load_factor = total_demand/max_demand
    if max_demand != 0:
        load_factor = total_demand / max_demand
    else:
        load_factor = -1

    # Writing data for "Maximum Power" fields in excel
    # Max wind section
    coordinates_for_maxWindGen = excel_column_name_to_coordinates["SystemSummary_MaxWindGen"]
    coordinates_for_maxWindGen_date = comms.increase_column_by_one(
        coordinates_for_maxWindGen)
    coordinates_for_maxWindGen_time = comms.increase_column_by_one(
        coordinates_for_maxWindGen_date)
    time_stamp_for_maxWindGen = wind_gen_to_time_stamp[max(list_of_wind_gen)]
    ws[coordinates_for_maxWindGen] = max(list_of_wind_gen)
    ws[coordinates_for_maxWindGen_date] = time_stamp_for_maxWindGen.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxWindGen_time] = time_stamp_for_maxWindGen.time().strftime('%X')

    # Max solar section
    coordinates_for_maxSolarGen = excel_column_name_to_coordinates["SystemSummary_MaxSolarGen"]
    coordinates_for_maxSolarGen_date = comms.increase_column_by_one(
        coordinates_for_maxSolarGen)
    coordinates_for_maxSolarGen_time = comms.increase_column_by_one(
        coordinates_for_maxSolarGen_date)
    time_stamp_for_maxSolarGen = solar_gen_to_time_stamp[max(
        list_of_solar_gen)]
    ws[coordinates_for_maxSolarGen] = max(list_of_solar_gen)
    ws[coordinates_for_maxSolarGen_date] = time_stamp_for_maxSolarGen.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxSolarGen_time] = time_stamp_for_maxSolarGen.time().strftime('%X')

    # Max AC generation section
    coordinates_for_maxTotalACGen = excel_column_name_to_coordinates["SystemSummary_MaxACGen"]
    coordinates_for_maxTotalACGen_date = comms.increase_column_by_one(
        coordinates_for_maxTotalACGen)
    coordinates_for_maxTotalACGen_time = comms.increase_column_by_one(
        coordinates_for_maxTotalACGen_date)
    time_stamp_for_maxTotalACGen = AC_total_gen_to_time_stamp[max(
        list_of_AC_total_gen)]
    ws[coordinates_for_maxTotalACGen] = max(list_of_AC_total_gen)
    ws[coordinates_for_maxTotalACGen_date] = time_stamp_for_maxTotalACGen.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxTotalACGen_time] = time_stamp_for_maxTotalACGen.time().strftime('%X')

    # Max DC generation section
    coordinates_for_maxTotalDCGen = excel_column_name_to_coordinates["SystemSummary_MaxDCGen"]
    coordinates_for_maxTotalDCGen_date = comms.increase_column_by_one(
        coordinates_for_maxTotalDCGen)
    coordinates_for_maxTotalDCGen_time = comms.increase_column_by_one(
        coordinates_for_maxTotalDCGen_date)
    time_stamp_for_maxTotalDCGen = DC_total_gen_to_time_stamp[max(
        list_of_DC_total_gen)]
    ws[coordinates_for_maxTotalDCGen] = max(list_of_DC_total_gen)
    ws[coordinates_for_maxTotalDCGen_date] = time_stamp_for_maxTotalDCGen.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxTotalDCGen_time] = time_stamp_for_maxTotalDCGen.time().strftime('%X')

    # Max AC Demand section
    coordinates_for_maxTotalACLoad = excel_column_name_to_coordinates["SystemSummary_MaxACLoad"]
    coordinates_for_maxTotalACLoad_date = comms.increase_column_by_one(
        coordinates_for_maxTotalACLoad)
    coordinates_for_maxTotalACLoad_time = comms.increase_column_by_one(
        coordinates_for_maxTotalACLoad_date)
    time_stamp_for_maxTotalACLoad = AC_total_load_to_time_stamp[max(
        list_of_AC_total_load)]
    ws[coordinates_for_maxTotalACLoad] = max(list_of_AC_total_load)
    ws[coordinates_for_maxTotalACLoad_date] = time_stamp_for_maxTotalACLoad.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxTotalACLoad_time] = time_stamp_for_maxTotalACLoad.time(
    ).strftime('%X')

    # Max DC Demand section
    coordinates_for_maxTotalDCLoad = excel_column_name_to_coordinates["SystemSummary_MaxDCLoad"]
    coordinates_for_maxTotalDCLoad_date = comms.increase_column_by_one(
        coordinates_for_maxTotalDCLoad)
    coordinates_for_maxTotalDCLoad_time = comms.increase_column_by_one(
        coordinates_for_maxTotalDCLoad_date)
    time_stamp_for_maxTotalDCLoad = DC_total_load_to_time_stamp[max(
        list_of_DC_total_load)]
    ws[coordinates_for_maxTotalDCLoad] = max(list_of_DC_total_load)
    ws[coordinates_for_maxTotalDCLoad_date] = time_stamp_for_maxTotalDCLoad.date(
    ).strftime('%m-%d-%Y')
    ws[coordinates_for_maxTotalDCLoad_time] = time_stamp_for_maxTotalDCLoad.time(
    ).strftime('%X')

    # Writing data for "Maximum Voltage" fields in excel
    # Max PhaseA Voltage section
    coordinates_for_maxVaBusID = excel_column_name_to_coordinates["SystemSummary_MaxPhaseA"]
    ws[coordinates_for_maxVaBusID] = Va_max_to_resultID_busID[max(
        Va_max_list)][1]
    coordinates_for_maxVaMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_maxVaBusID))
    ws[coordinates_for_maxVaMag] = max(Va_max_list)
    time_stamp_maxVa = resultID_to_timestamp_dict[Va_max_to_resultID_busID[max(
        Va_max_list)][0]][0]
    date_value_maxVaDate, time_value_maxVaTime = comms.date_and_time_from_timeStamp(
        time_stamp_maxVa)
    coordinates_for_maxVaDate = comms.increase_column_by_one(
        coordinates_for_maxVaMag)
    ws[coordinates_for_maxVaDate] = date_value_maxVaDate.strftime('%m-%d-%Y')
    coordinates_for_maxVaTime = comms.increase_column_by_one(
        coordinates_for_maxVaDate)
    ws[coordinates_for_maxVaTime] = time_value_maxVaTime

    # Max PhaseB Voltage section
    coordinates_for_maxVbBusID = excel_column_name_to_coordinates["SystemSummary_MaxPhaseB"]
    ws[coordinates_for_maxVbBusID] = Vb_max_to_resultID_busID[max(
        Vb_max_list)][1]
    coordinates_for_maxVbMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_maxVbBusID))
    ws[coordinates_for_maxVbMag] = max(Vb_max_list)
    time_stamp_maxVb = resultID_to_timestamp_dict[Vb_max_to_resultID_busID[max(
        Vb_max_list)][0]][0]
    date_value_maxVbDate, time_value_maxVbTime = comms.date_and_time_from_timeStamp(
        time_stamp_maxVb)
    coordinates_for_maxVbDate = comms.increase_column_by_one(
        coordinates_for_maxVbMag)
    ws[coordinates_for_maxVbDate] = date_value_maxVbDate.strftime('%m-%d-%Y')
    coordinates_for_maxVbTime = comms.increase_column_by_one(
        coordinates_for_maxVbDate)
    ws[coordinates_for_maxVbTime] = time_value_maxVbTime

    # Max PhaseC Voltage section
    coordinates_for_maxVcBusID = excel_column_name_to_coordinates["SystemSummary_MaxPhaseC"]
    ws[coordinates_for_maxVcBusID] = Vc_max_to_resultID_busID[max(
        Vc_max_list)][1]
    coordinates_for_maxVcMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_maxVcBusID))
    ws[coordinates_for_maxVcMag] = max(Vc_max_list)
    time_stamp_maxVc = resultID_to_timestamp_dict[Vc_max_to_resultID_busID[max(
        Vc_max_list)][0]][0]
    date_value_maxVcDate, time_value_maxVcTime = comms.date_and_time_from_timeStamp(
        time_stamp_maxVc)
    coordinates_for_maxVcDate = comms.increase_column_by_one(
        coordinates_for_maxVcMag)
    ws[coordinates_for_maxVcDate] = date_value_maxVcDate.strftime('%m-%d-%Y')
    coordinates_for_maxVcTime = comms.increase_column_by_one(
        coordinates_for_maxVcDate)
    ws[coordinates_for_maxVcTime] = time_value_maxVcTime

    # Writing data for "Minimum Voltage" fields in excel
    # Min PhaseA Voltage section
    coordinates_for_minVaBusID = excel_column_name_to_coordinates["SystemSummary_MinPhaseA"]
    ws[coordinates_for_minVaBusID] = Va_min_to_resultID_busID[min(
        Va_min_list)][1]
    coordinates_for_minVaMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_minVaBusID))
    ws[coordinates_for_minVaMag] = max(Va_min_list)
    time_stamp_minVa = resultID_to_timestamp_dict[Va_min_to_resultID_busID[min(
        Va_min_list)][0]][0]
    date_value_minVaDate, time_value_minVaTime = comms.date_and_time_from_timeStamp(
        time_stamp_minVa)
    coordinates_for_minVaDate = comms.increase_column_by_one(
        coordinates_for_minVaMag)
    ws[coordinates_for_minVaDate] = date_value_minVaDate.strftime('%m-%d-%Y')
    coordinates_for_minVaTime = comms.increase_column_by_one(
        coordinates_for_minVaDate)
    ws[coordinates_for_minVaTime] = time_value_minVaTime

    # Min PhaseB Voltage section
    coordinates_for_minVbBusID = excel_column_name_to_coordinates["SystemSummary_MinPhaseB"]
    ws[coordinates_for_minVbBusID] = Vb_min_to_resultID_busID[min(
        Vb_min_list)][1]
    coordinates_for_minVbMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_minVbBusID))
    ws[coordinates_for_minVbMag] = max(Vb_min_list)
    time_stamp_minVb = resultID_to_timestamp_dict[Vb_min_to_resultID_busID[min(
        Vb_min_list)][0]][0]
    date_value_minVbDate, time_value_minVbTime = comms.date_and_time_from_timeStamp(
        time_stamp_minVb)
    coordinates_for_minVbDate = comms.increase_column_by_one(
        coordinates_for_minVbMag)
    ws[coordinates_for_minVbDate] = date_value_minVbDate.strftime('%m-%d-%Y')
    coordinates_for_minVbTime = comms.increase_column_by_one(
        coordinates_for_minVbDate)
    ws[coordinates_for_minVbTime] = time_value_minVbTime

    # Min PhaseC Voltage section
    coordinates_for_minVcBusID = excel_column_name_to_coordinates["SystemSummary_MinPhaseC"]
    ws[coordinates_for_minVcBusID] = Vc_min_to_resultID_busID[max(
        Vc_min_list)][1]
    coordinates_for_minVcMag = comms.increase_column_by_one(
        comms.increase_column_by_one(coordinates_for_minVcBusID))
    ws[coordinates_for_minVcMag] = max(Vc_min_list)
    time_stamp_minVc = resultID_to_timestamp_dict[Vc_min_to_resultID_busID[min(
        Vc_min_list)][0]][0]
    date_value_minVcDate, time_value_minVcTime = comms.date_and_time_from_timeStamp(
        time_stamp_minVc)
    coordinates_for_minVcDate = comms.increase_column_by_one(
        coordinates_for_minVcMag)
    ws[coordinates_for_minVcDate] = date_value_minVcDate.strftime('%m-%d-%Y')
    coordinates_for_minVcTime = comms.increase_column_by_one(
        coordinates_for_minVcDate)
    ws[coordinates_for_minVcTime] = time_value_minVcTime

    # Writing calculated data for "Total Enegry" fields in excel
    coordinates_for_windGen = excel_column_name_to_coordinates["SystemSummary_TotalWindGen"]
    ws[coordinates_for_windGen] = total_wind_energy

    coordinates_for_solarGen = excel_column_name_to_coordinates["SystemSummary_TotalSolarGen"]
    ws[coordinates_for_solarGen] = total_solar_energy

    coordinates_for_ACGen = excel_column_name_to_coordinates["SystemSummary_TotalACGen"]
    ws[coordinates_for_ACGen] = total_AC_generation_energy

    coordinates_for_DCGen = excel_column_name_to_coordinates["SystemSummary_TotalDCGen"]
    ws[coordinates_for_DCGen] = total_DC_generation_energy

    coordinates_for_sysGen = excel_column_name_to_coordinates["SystemSummary_TotalSysGen"]
    ws[coordinates_for_sysGen] = total_system_generation_energy

    coordinates_for_ACLoad = excel_column_name_to_coordinates["SystemSummary_TotalACLoad"]
    ws[coordinates_for_ACLoad] = total_AC_load_energy

    coordinates_for_DCLoad = excel_column_name_to_coordinates["SystemSummary_TotalDCLoad"]
    ws[coordinates_for_DCLoad] = total_DC_load_energy

    coordinates_for_sysLoad = excel_column_name_to_coordinates["SystemSummary_TotalSysLoad"]
    ws[coordinates_for_sysLoad] = total_system_load_energy

    coordinates_for_loss = excel_column_name_to_coordinates["SystemSummary_TotalLoss"]
    ws[coordinates_for_loss] = total_energy_loss

    coordinates_for_loadfactor = excel_column_name_to_coordinates["SystemSummary_LoadFactor"]
    # ws[coordinates_for_loadfactor] = load_factor

    if load_factor != -1:
        ws[coordinates_for_loadfactor] = load_factor
    else:
        ws[coordinates_for_loadfactor] = 'N/A'


def load_cover_data(wb, ws, cursor, table_names_list, templatePath):
    # comms.check_db_version(table_names_list, cursor)

    # Getting train study case info table
    # table_index = table_names_list.index('tdstudycaseinfo')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDStudyCaseInfo"
    cursor.execute(command)
    study_case_info_data = cursor.fetchall()
    total_rows_study_case_info = len(study_case_info_data)
    study_case_info_column_name_list = [
        description[0] for description in cursor.description]

    # Getting train header info table
    # table_index = table_names_list.index('headr')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM Headr"
    cursor.execute(command)
    header_info_data = cursor.fetchall()
    total_rows_header_info = len(header_info_data)
    header_info_column_name_list = [description[0]
                                    for description in cursor.description]

    column_names = len(wb.defined_names.definedName)
    defined_names_for_project_info = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] == "Project":
            defined_names_for_project_info.append(
                wb.defined_names.definedName[i].name)

    # nameActive = openpyxl.writer.workbook.get_active_sheet(wb)

    excel_column_name_to_coordinates_info = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, len(defined_names_for_project_info)):
        try:
            info_excel_column_name = defined_names_for_project_info[i]
            info_actual_coordinates = next(
                wb.defined_names[info_excel_column_name].destinations)[1]
            # ++[JaykumarD] 2018/07/13 IR-53983 - TDULF, Report, 3 issues found with Excel report, ETAP version headr and adjustment info are wrong
            #info_start_writing_coordinate = comms.increase_column_by_one(info_actual_coordinates)
            if info_excel_column_name == "Cover_PSRev":
                info_start_writing_coordinate = comms.get_updated_coordinates(
                    info_actual_coordinates)
            else:
                info_start_writing_coordinate = comms.increase_column_by_one(
                    info_actual_coordinates)
            excel_column_name_to_coordinates_info[info_excel_column_name] = info_start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    # Writing data for all the fields from header table in excel
    coordinates_for_projectName = excel_column_name_to_coordinates_info["Cover_Project"]
    ws[coordinates_for_projectName] = header_info_data[0][0]

    coordinates_for_PSRev = excel_column_name_to_coordinates_info["Cover_PSRev"]
    ws[coordinates_for_PSRev] = header_info_data[0][4]

    coordinates_for_Loc = excel_column_name_to_coordinates_info["Cover_Loc"]
    ws[coordinates_for_Loc] = header_info_data[0][1]

    coordinates_for_Contract = excel_column_name_to_coordinates_info["Cover_Contr"]
    ws[coordinates_for_Contract] = header_info_data[0][2]

    coordinates_for_Date = excel_column_name_to_coordinates_info["Cover_Date"]
    ws[coordinates_for_Date] = header_info_data[0][6]

    coordinates_for_Engineer = excel_column_name_to_coordinates_info["Cover_Eng"]
    ws[coordinates_for_Engineer] = header_info_data[0][3]

    coordinates_for_projectFileName = excel_column_name_to_coordinates_info["Cover_FileN"]
    ws[coordinates_for_projectFileName] = header_info_data[0][8]

    coordinates_for_serialNumber = excel_column_name_to_coordinates_info["Cover_SN"]
    ws[coordinates_for_serialNumber] = header_info_data[0][7]

    # Writing data for all the fields from StudyCase Info table in excel
    coordinates_for_studyCaseName = excel_column_name_to_coordinates_info["Cover_STDCase"]
    ws[coordinates_for_studyCaseName] = study_case_info_data[0][12]

    coordinates_for_OutputFileName = excel_column_name_to_coordinates_info["Cover_Output"]
    ws[coordinates_for_OutputFileName] = study_case_info_data[0][30]

    coordinates_for_config = excel_column_name_to_coordinates_info["Cover_config"]
    ws[coordinates_for_config] = study_case_info_data[0][31]

    coordinates_for_revision = excel_column_name_to_coordinates_info["Cover_Revision"]
    ws[coordinates_for_revision] = study_case_info_data[0][32]

    coordinates_for_standard = excel_column_name_to_coordinates_info["Cover_Standard"]
    ws[coordinates_for_standard] = study_case_info_data[0][28]

    coordinates_for_freq = excel_column_name_to_coordinates_info["Cover_Freq"]
    ws[coordinates_for_freq] = study_case_info_data[0][26]

    coordinates_for_unitSystem = excel_column_name_to_coordinates_info["Cover_Unit"]
    ws[coordinates_for_unitSystem] = study_case_info_data[0][27]

    path_name = os.path.dirname(templatePath)
    path_name2 = os.path.dirname(path_name)
    imageLocation = path_name2 + r'\etap_logo.png'

    img = openpyxl.drawing.image.Image(imageLocation)
    ws.add_image(img, 'E4')
    ws.sheet_view.showGridLines = False
    comms.patch_worksheet()


def load_study_info_data(wb, ws, cursor, table_names_list):
    # comms.check_db_version(table_names_list, cursor)

    # Getting TDLF Study Case info table
    # table_index = table_names_list.index('tdstudycaseinfo')
    # command = "SELECT * FROM " + table_names_list[table_index]
    command = "SELECT * FROM TDStudyCaseInfo"
    cursor.execute(command)
    study_case_info_data = cursor.fetchall()
    total_rows_study_case_info = len(study_case_info_data)
    study_case_info_column_name_list = [
        description[0] for description in cursor.description]

    column_names = len(wb.defined_names.definedName)
    defined_names_for_project_info = []

    for i in range(0, column_names):
        if next(wb.defined_names.definedName[i].destinations)[0] == "Info":
            defined_names_for_project_info.append(
                wb.defined_names.definedName[i].name)

    # nameActive = openpyxl.writer.workbook.get_active_sheet(wb)

    excel_column_name_to_coordinates_info = dict()
    # Find the current coordinates and update the starting coordinates
    for i in range(0, len(defined_names_for_project_info)):
        try:
            info_excel_column_name = defined_names_for_project_info[i]
            if (info_excel_column_name == "Info_ConstP" or info_excel_column_name == "Info_ConstZ" or info_excel_column_name == "Info_ConstI" or info_excel_column_name == "Info_ConstGen"):
                info_actual_coordinates = next(
                    wb.defined_names[info_excel_column_name].destinations)[1]
                info_start_writing_coordinate = comms.get_updated_coordinates(
                    info_actual_coordinates)
                excel_column_name_to_coordinates_info[info_excel_column_name] = info_start_writing_coordinate
            else:
                info_actual_coordinates = next(
                    wb.defined_names[info_excel_column_name].destinations)[1]
                info_start_writing_coordinate = comms.increase_column_by_one(comms.increase_column_by_one(
                    comms.increase_column_by_one(comms.increase_column_by_one(comms.increase_column_by_one(info_actual_coordinates)))))
                excel_column_name_to_coordinates_info[info_excel_column_name] = info_start_writing_coordinate
        except KeyError:
            errors = '1'
            continue

    # Writing data for all the fields from TD study case table in excel
    coordinates_for_studyCaseID = excel_column_name_to_coordinates_info["Info_StudyCaseID"]
    ws[coordinates_for_studyCaseID] = study_case_info_data[0][12]

    coordinates_for_maxIteration = excel_column_name_to_coordinates_info["Info_MaxIteration"]
    ws[coordinates_for_maxIteration] = study_case_info_data[0][13]

    coordinates_for_solPrecision = excel_column_name_to_coordinates_info[
        "Info_SolutionPrecision"]
    ws[coordinates_for_solPrecision] = study_case_info_data[0][14]

    coordinates_for_haltNonConverge = excel_column_name_to_coordinates_info[
        "Info_HaltNonConverge"]
    ws[coordinates_for_haltNonConverge] = comms.is_tolerance_applied(
        study_case_info_data[0][15])

    coordinates_for_haltEquipOverLoad = excel_column_name_to_coordinates_info[
        "Info_HaltEquipOverLoad"]
    ws[coordinates_for_haltEquipOverLoad] = comms.is_tolerance_applied(
        study_case_info_data[0][16])

    coordinates_for_aCDCSimSolution = excel_column_name_to_coordinates_info[
        "Info_ACDCSimSolution"]
    ws[coordinates_for_aCDCSimSolution] = comms.is_tolerance_applied(
        study_case_info_data[0][17])

    coordinates_for_loadingCat = excel_column_name_to_coordinates_info["Info_LoadingCat"]
    ws[coordinates_for_loadingCat] = study_case_info_data[0][18]

    coordinates_for_genCat = excel_column_name_to_coordinates_info["Info_GenCat"]
    ws[coordinates_for_genCat] = study_case_info_data[0][20]

    coordinates_for_dCLoadingCat = excel_column_name_to_coordinates_info["Info_DCLoadingCat"]
    ws[coordinates_for_dCLoadingCat] = study_case_info_data[0][22]

    charger_loading_final_raw = study_case_info_data[0][24]
    if charger_loading_final_raw == 1:
        charger_loading_final = "Loading Category"
    elif charger_loading_final_raw == 2:
        charger_loading_final = "Operating Load"
    elif charger_loading_final_raw == 3:
        charger_loading_final = "Constant Current"
    elif charger_loading_final_raw == 4:
        charger_loading_final = "Uniform"
    else:
        charger_loading_final = "NA"
    coordinates_for_chargerLoading = excel_column_name_to_coordinates_info[
        "Info_ChargerLoading"]
    ws[coordinates_for_chargerLoading] = charger_loading_final

    load_diversity_factor_raw = study_case_info_data[0][25]
    if load_diversity_factor_raw == 0:
        ldf = "None"
    elif load_diversity_factor_raw == 1:
        ldf = "Bus Maximum"
    elif load_diversity_factor_raw == 2:
        ldf = "Bus Minimum"
    elif load_diversity_factor_raw == 3:
        ldf = "Global"
    else:
        ldf = "NA"
    coordinates_for_lDFactor = excel_column_name_to_coordinates_info["Info_LDFactor"]
    ws[coordinates_for_lDFactor] = ldf

    coordinates_for_constP = excel_column_name_to_coordinates_info["Info_ConstP"]
    ws[coordinates_for_constP] = study_case_info_data[0][33]

    coordinates_for_constZ = excel_column_name_to_coordinates_info["Info_ConstZ"]
    ws[coordinates_for_constZ] = study_case_info_data[0][34]

    coordinates_for_constI = excel_column_name_to_coordinates_info["Info_ConstI"]
    ws[coordinates_for_constI] = study_case_info_data[0][35]

    coordinates_for_constGen = excel_column_name_to_coordinates_info["Info_ConstGen"]
    ws[coordinates_for_constGen] = study_case_info_data[0][36]

    simulation_start_time_stamp = study_case_info_data[0][1]
    simulation_start_date = simulation_start_time_stamp[:simulation_start_time_stamp.find(
        ' ')]
    simulation_start_time = simulation_start_time_stamp[simulation_start_time_stamp.find(
        ' ')+1:]
    simulation_start_date_final = datetime.strptime(
        simulation_start_date, '%m-%d-%Y').date()

    simulation_stop_time_stamp = study_case_info_data[0][2]
    simulation_stop_date = simulation_stop_time_stamp[:simulation_stop_time_stamp.find(
        ' ')]
    simulation_stop_time = simulation_stop_time_stamp[simulation_stop_time_stamp.find(
        ' ')+1:]
    simulation_stop_date_final = datetime.strptime(
        simulation_stop_date, '%m-%d-%Y').date()

    coordinates_for_simStartDate = excel_column_name_to_coordinates_info["Info_SimStartTime"]
    ws[coordinates_for_simStartDate] = simulation_start_date

    coordinates_for_simStartTime = comms.increase_column_by_one(
        coordinates_for_simStartDate)
    ws[coordinates_for_simStartTime] = simulation_start_time

    coordinates_for_simStopDate = excel_column_name_to_coordinates_info["Info_SimStopTime"]
    ws[coordinates_for_simStopDate] = simulation_stop_date

    coordinates_for_simStartTime = comms.increase_column_by_one(
        coordinates_for_simStopDate)
    ws[coordinates_for_simStartTime] = simulation_stop_time

    coordinates_for_simStepValue = excel_column_name_to_coordinates_info["Info_SimStep"]
    ws[coordinates_for_simStepValue] = study_case_info_data[0][3]

    coordinates_for_simStepUnit = comms.increase_column_by_one(
        coordinates_for_simStepValue)
    ws[coordinates_for_simStepUnit] = study_case_info_data[0][10]

    coordinates_for_totalSimCount = excel_column_name_to_coordinates_info["Info_TotalSimCount"]
    ws[coordinates_for_totalSimCount] = study_case_info_data[0][11]

    coordinates_for_externalData = excel_column_name_to_coordinates_info["Info_ExternalData"]
    ws[coordinates_for_externalData] = comms.is_tolerance_applied(
        study_case_info_data[0][39])

    wind_external_data_raw = study_case_info_data[0][40]
    if wind_external_data_raw == 1:
        wind_external_data_final = "Wind"
    elif wind_external_data_raw == 2:
        wind_external_data_final = "PQ Data"
    else:
        wind_external_data_final = "NA"
    coordinates_for_windExternalData = excel_column_name_to_coordinates_info[
        "Info_WindExternalData"]
    ws[coordinates_for_windExternalData] = wind_external_data_final

    pv_external_data_raw = study_case_info_data[0][41]
    if pv_external_data_raw == 1:
        pv_external_data_final = "Iradiance"
    elif pv_external_data_raw == 2:
        pv_external_data_final = "PQ Data"
    else:
        pv_external_data_final = "NA"
    coordinates_for_pVExternalData = excel_column_name_to_coordinates_info[
        "Info_PVExternalData"]
    ws[coordinates_for_pVExternalData] = pv_external_data_final

    coordinates_for_annualLoadGrowth = excel_column_name_to_coordinates_info[
        "Info_AnnualLoadGrowthFactor"]
    ws[coordinates_for_annualLoadGrowth] = study_case_info_data[0][42]

    coordinates_for_applyNegTolMinTemp = excel_column_name_to_coordinates_info[
        "Info_ApplyNegTolMinTemp"]
    # ++[JaykumarD] 2018/07/13 IR-53983 - TDULF, Report, 3 issues found with Excel report, ETAP version headr and adjustment info are wrong
    #ws[coordinates_for_applyNegTolMinTemp] = comms.is_tolerance_applied(study_case_info_data[0][43])
    ws[coordinates_for_applyNegTolMinTemp] = comms.is_tolerance_applied(
        study_case_info_data[0][64])

    coordinates_for_xmerZTol = excel_column_name_to_coordinates_info["Info_XmerZTol"]
    ws[coordinates_for_xmerZTol] = comms.is_tolerance_applied(
        study_case_info_data[0][43])

    coordinates_for_xmerZTolIndGob = comms.increase_column_by_one(
        coordinates_for_xmerZTol)
    ws[coordinates_for_xmerZTolIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][44])

    coordinates_for_xmerZTolPercent = comms.increase_column_by_one(
        coordinates_for_xmerZTolIndGob)
    ws[coordinates_for_xmerZTolPercent] = study_case_info_data[0][45]

    coordinates_for_xmerXTol = excel_column_name_to_coordinates_info["Info_XmerXTol"]
    ws[coordinates_for_xmerXTol] = comms.is_tolerance_applied(
        study_case_info_data[0][61])

    coordinates_for_xmerXTolIndGob = comms.increase_column_by_one(
        coordinates_for_xmerXTol)
    ws[coordinates_for_xmerXTolIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][62])

    coordinates_for_xmerXTolPercent = comms.increase_column_by_one(
        coordinates_for_xmerXTolIndGob)
    ws[coordinates_for_xmerXTolPercent] = study_case_info_data[0][63]

    coordinates_for_oLHResistance = excel_column_name_to_coordinates_info["Info_OLHResistance"]
    ws[coordinates_for_oLHResistance] = comms.is_tolerance_applied(
        study_case_info_data[0][52])

    coordinates_for_oLHResistanceIndGob = comms.increase_column_by_one(
        coordinates_for_oLHResistance)
    ws[coordinates_for_oLHResistanceIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][53])

    coordinates_for_oLHResistancePercent = comms.increase_column_by_one(
        coordinates_for_oLHResistanceIndGob)
    ws[coordinates_for_oLHResistancePercent] = study_case_info_data[0][54]

    coordinates_for_lineLength = excel_column_name_to_coordinates_info["Info_LineLength"]
    ws[coordinates_for_lineLength] = comms.is_tolerance_applied(
        study_case_info_data[0][46])

    coordinates_for_lineLengthIndGob = comms.increase_column_by_one(
        coordinates_for_lineLength)
    ws[coordinates_for_lineLengthIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][47])

    coordinates_for_lineLengthPercent = comms.increase_column_by_one(
        coordinates_for_lineLengthIndGob)
    ws[coordinates_for_lineLengthPercent] = study_case_info_data[0][48]

    coordinates_for_cableLength = excel_column_name_to_coordinates_info["Info_CableLength"]
    ws[coordinates_for_cableLength] = comms.is_tolerance_applied(
        study_case_info_data[0][49])

    coordinates_for_cableLengthIndGob = comms.increase_column_by_one(
        coordinates_for_cableLength)
    ws[coordinates_for_cableLengthIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][50])

    coordinates_for_cableLengthPercent = comms.increase_column_by_one(
        coordinates_for_cableLengthIndGob)
    ws[coordinates_for_cableLengthPercent] = study_case_info_data[0][51]

    coordinates_for_lineResistance = excel_column_name_to_coordinates_info[
        "Info_LineResistance"]
    ws[coordinates_for_lineResistance] = comms.is_tolerance_applied(
        study_case_info_data[0][55])

    coordinates_for_lineResistanceIndGob = comms.increase_column_by_one(
        coordinates_for_lineResistance)
    ws[coordinates_for_lineResistanceIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][56])

    coordinates_for_lineResistancePercent = comms.increase_column_by_one(
        coordinates_for_lineResistanceIndGob)
    ws[coordinates_for_lineResistancePercent] = study_case_info_data[0][57]

    coordinates_for_cableResistance = excel_column_name_to_coordinates_info[
        "Info_CableResistance"]
    ws[coordinates_for_cableResistance] = comms.is_tolerance_applied(
        study_case_info_data[0][58])

    coordinates_for_cableResistanceIndGob = comms.increase_column_by_one(
        coordinates_for_cableResistance)
    ws[coordinates_for_cableResistanceIndGob] = comms.apply_global_or_indivudual_tolerance(
        study_case_info_data[0][59])

    coordinates_for_cableResistancePercent = comms.increase_column_by_one(
        coordinates_for_cableResistanceIndGob)
    ws[coordinates_for_cableResistancePercent] = study_case_info_data[0][60]

    coordinates_for_calcNumber = excel_column_name_to_coordinates_info["Info_CalcNumber"]
    ws[coordinates_for_calcNumber] = study_case_info_data[0][5]

    coordinates_for_nonConvCalc = excel_column_name_to_coordinates_info["Info_NonConvCalc"]
    ws[coordinates_for_nonConvCalc] = study_case_info_data[0][6]

    coordinates_for_haltingCalcUnsolved = excel_column_name_to_coordinates_info[
        "Info_HaltingCalcUnsolved"]
    ws[coordinates_for_haltingCalcUnsolved] = study_case_info_data[0][7]

    calcStart_time_stamp = study_case_info_data[0][8]
    calcStart_date_value = calcStart_time_stamp[:calcStart_time_stamp.find(
        ' ')]
    calcStart_time_value = calcStart_time_stamp[calcStart_time_stamp.find(
        ' ')+1:]
    calcStart_date_final = datetime.strptime(
        calcStart_date_value, '%m-%d-%Y').date()

    calcStop_time_stamp = study_case_info_data[0][9]
    calcStop_date_value = calcStop_time_stamp[:calcStop_time_stamp.find(' ')]
    calcStop_time_value = calcStop_time_stamp[calcStop_time_stamp.find(' ')+1:]
    calcStop_date_final = datetime.strptime(
        calcStop_date_value, '%m-%d-%Y').date()

    coordinates_for_calcStartDate = excel_column_name_to_coordinates_info["Info_CalcStartDate"]
    ws[coordinates_for_calcStartDate] = calcStart_date_value

    coordinates_for_calcStartTime = comms.increase_column_by_one(
        coordinates_for_calcStartDate)
    ws[coordinates_for_calcStartTime] = calcStart_time_value

    coordinates_for_calcStopDate = excel_column_name_to_coordinates_info["Info_CalcStopDate"]
    ws[coordinates_for_calcStopDate] = calcStop_date_value

    coordinates_for_calcStopTime = comms.increase_column_by_one(
        coordinates_for_calcStopDate)
    ws[coordinates_for_calcStopTime] = calcStop_time_value

    coordinates_for_battery_model = \
        comms.get_updated_columns_coordinates(
            excel_column_name_to_coordinates_info["Battery_Model"], 25, -4)
    if study_case_info_data[0][65] == 0:
        ws[coordinates_for_battery_model] = r'Rated Voc & R'
    elif study_case_info_data[0][65] == 1:
        ws[coordinates_for_battery_model] = 'SOC Category'

        coordinates_for_soc_cat = \
            comms.get_updated_columns_coordinates(
                excel_column_name_to_coordinates_info["SOC_Cat"], 26, -4)
        ws[coordinates_for_soc_cat] = study_case_info_data[0][66]

        coordinates_for_soc_value = \
            comms.get_updated_columns_coordinates(
                excel_column_name_to_coordinates_info["SOC_value"], 27, -4)
        ws[coordinates_for_soc_value] = study_case_info_data[0][67]
    else:
        ws[coordinates_for_battery_model] = 'Global SOC'
        coordinates_for_soc_value = \
            comms.get_updated_columns_coordinates(
                excel_column_name_to_coordinates_info["SOC_value"], 27, -4)
        ws[coordinates_for_soc_value] = study_case_info_data[0][67]


def load_energy_storage_data(wb, ws, cursor, table_names_list):
    command = '''select tdtimeid.timeid, tdtimeid.time, tdenergystorage.deviceid, tdenergystorage.termv, 
    tdenergystorage.ampflow, tdenergystorage.powerflow, tdenergystorage.energystored, tdenergystorage.soc
    from tdenergystorage join tdtimeid
    on tdtimeid.resultid = tdenergystorage.resultid
    order by tdenergystorage.deviceid, tdtimeid.timeid
    '''
    cursor.execute(command)
    data = cursor.fetchall()
    result = dict()
    coordinates_dict = comms.name_to_coordinate(wb)
    row_style = comms.get_row_style(wb, ws, coordinates_dict)

    small_value = 0.0001  # for determining charging or discharging status for energystored

    command_time_step = 'select simulationhoursperstep, simulationstepunit from tdstudycaseinfo'
    cursor.execute(command_time_step)
    simulation_time_step = cursor.fetchall()
    if simulation_time_step[0][1] == 'Minute':
        simulation_time_hour = float(simulation_time_step[0][0]) / 60
    elif simulation_time_step[0][1] == 'Second':
        simulation_time_hour = float(simulation_time_step[0][0]) / 3600
    elif simulation_time_step[0][1] == 'Day':
        simulation_time_hour = float(simulation_time_step[0][0]) * 24
    elif simulation_time_step[0][1] == 'Week':
        simulation_time_hour = float(simulation_time_step[0][0]) * 24 * 7
    elif simulation_time_step[0][1] == 'Month':
        simulation_time_hour = float(simulation_time_step[0][0]) * 24 * 7 * 30
    elif simulation_time_step[0][1] == 'Single':
        simulation_time_hour = 0
    else:
        simulation_time_hour = float(simulation_time_step[0][0])
    for i in range(len(data)):
        if data[i][2] not in result.keys():
            temp_last_soc = data[i][7]
            temp_delta_soc = 0
            temp_count = 1
            temp_total_soc = temp_last_soc
            if data[i][6] >= small_value:  # energystored >= 0  charging
                if simulation_time_step[0][1] != 'Single':
                    temp_total_charge_energy = data[i][6]
                else:
                    temp_total_charge_energy = 0
                temp_total_charge_time = 1 * simulation_time_hour
                temp_charge_max_P = abs(data[i][5])
                temp_total_discharge_energy = 0
                temp_total_discharge_time = 0
                temp_discharge_max_P = 0
            elif data[i][6] <= -small_value:
                temp_total_charge_energy = 0
                temp_total_charge_time = 0
                temp_charge_max_P = 0
                if simulation_time_step[0][1] != 'Single':
                    temp_total_discharge_energy = abs(data[i][6])
                else:
                    temp_total_discharge_energy = 0
                temp_total_discharge_time = 1 * simulation_time_hour
                temp_discharge_max_P = abs(data[i][5])
            else:
                temp_total_charge_energy = 0
                temp_total_charge_time = 0
                temp_charge_max_P = 0
                temp_total_discharge_energy = 0
                temp_total_discharge_time = 0
                temp_discharge_max_P = 0

            result[data[i][2]] = [temp_total_charge_energy, temp_total_charge_time, temp_total_discharge_energy,
                                  temp_total_discharge_time, temp_charge_max_P, temp_discharge_max_P,
                                  float(temp_delta_soc/100),
                                  float(temp_total_soc/temp_count)]
        else:
            temp_current_soc = data[i][7]
            temp_total_soc += temp_current_soc
            temp_count += 1
            temp_delta_soc += float(abs(temp_current_soc - temp_last_soc)/100)
            temp_last_soc = temp_current_soc
            if data[i][6] >= small_value:  # energystored >= 0  charging
                temp_total_charge_energy += data[i][6]
                temp_total_charge_time += 1 * simulation_time_hour
                if abs(data[i][5]) > temp_charge_max_P:
                    temp_charge_max_P = abs(data[i][5])
            elif data[i][6] <= -small_value:
                temp_total_discharge_energy += abs(data[i][6])
                temp_total_discharge_time += 1 * simulation_time_hour
                if abs(data[i][5]) > temp_discharge_max_P:
                    temp_discharge_max_P = abs(data[i][5])
            else:
                pass
            result[data[i][2]] = [temp_total_charge_energy, temp_total_charge_time, temp_total_discharge_energy,
                                  temp_total_discharge_time, temp_charge_max_P, temp_discharge_max_P, temp_delta_soc,
                                  float(temp_total_soc / temp_count)]
    for key, value in result.items():
        ws[coordinates_dict['Battery_ID'][1]]._style = row_style['Battery_ID']
        ws[coordinates_dict['Battery_ID'][1]].value = key
        coordinates_dict['Battery_ID'][1] = comms.get_updated_coordinates_inc(
            coordinates_dict['Battery_ID'][1], 1)

        ws[coordinates_dict['Energy_Charge'][1]
           ]._style = row_style['Energy_Charge']
        ws[coordinates_dict['Energy_Charge'][1]].value = value[0]
        coordinates_dict['Energy_Charge'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['Energy_Charge'][1], 1)

        ws[coordinates_dict['Time_Charge'][1]]._style = row_style['Time_Charge']
        ws[coordinates_dict['Time_Charge'][1]].value = value[1]
        coordinates_dict['Time_Charge'][1] = comms.get_updated_coordinates_inc(
            coordinates_dict['Time_Charge'][1], 1)

        ws[coordinates_dict['Energy_Discharge'][1]
           ]._style = row_style['Energy_Discharge']
        ws[coordinates_dict['Energy_Discharge'][1]].value = value[2]
        coordinates_dict['Energy_Discharge'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['Energy_Discharge'][1], 1)

        ws[coordinates_dict['Time_Discharge'][1]
           ]._style = row_style['Time_Discharge']
        ws[coordinates_dict['Time_Discharge'][1]].value = value[3]
        coordinates_dict['Time_Discharge'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['Time_Discharge'][1], 1)

        ws[coordinates_dict['MaxP_Charge'][1]]._style = row_style['MaxP_Charge']
        ws[coordinates_dict['MaxP_Charge'][1]].value = value[4]
        coordinates_dict['MaxP_Charge'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['MaxP_Charge'][1], 1)

        ws[coordinates_dict['MaxP_Discharge'][1]
           ]._style = row_style['MaxP_Discharge']
        ws[coordinates_dict['MaxP_Discharge'][1]].value = value[5]
        coordinates_dict['MaxP_Discharge'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['MaxP_Discharge'][1], 1)

        ws[coordinates_dict['Total_Cycle'][1]]._style = row_style['Total_Cycle']
        ws[coordinates_dict['Total_Cycle'][1]].value = value[6] * 100
        coordinates_dict['Total_Cycle'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['Total_Cycle'][1], 1)

        ws[coordinates_dict['Average_SOC'][1]]._style = row_style['Average_SOC']
        ws[coordinates_dict['Average_SOC'][1]].value = value[7]
        coordinates_dict['Average_SOC'][1] = \
            comms.get_updated_coordinates_inc(
                coordinates_dict['Average_SOC'][1], 1)
